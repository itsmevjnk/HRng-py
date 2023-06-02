#!/usr/bin/env python
# coding: utf-8
# cython: language_level=3

# CODE (SẮP) THƯƠNG MẠI HOÁ - VUI LÒNG KHÔNG LEAK!
# Nếu muốn public mã nguồn, vui lòng liên hệ tác giả tại:
#   itsmevjnk.work@gmail.com
#     hoặc
#   113ly.vinh.nguyenthanh@c3chuvanan.edu.vn

# và một lần nữa chúng ta lại quay về thời selenium
# prerequisites: pip install requests lxml pyqt5 pycryptodome python-dateutil selenium dateparser pyperclip openpyxl cython setuptools

#
#                       _oo0oo_
#                      o8888888o
#                      88" . "88
#                      (| -_- |)
#                      0\  =  /0
#                    ___/`---'\___
#                  .' \\|     |// '.
#                 / \\|||  :  |||// \
#                / _||||| -:- |||||- \
#               |   | \\\  -  /// |   |
#               | \_|  ''\---/''  |_/ |
#               \  .-\__  '-'  ___/-. /
#             ___'. .'  /--.--\  `. .'___
#          ."" '<  `.___\_<|>_/___.' >' "".
#         | | :  `- \`.;`\ _ /`;.`/ - ` : | |
#         \  \ `_.   \_ __\ /__ _/   .-` /  /
#     =====`-.____`.___ \_____/___.-`___.-'=====
#                       `=---='
#
#
#     ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#
#               佛祖保佑         永无BUG
#

import argparse
import sys

# thông tin phiên bản
sw_name = "HRng" # tên
sw_rel = 5 # số hiệu phiên bản (phải là số tự nhiên)

parser = argparse.ArgumentParser(description=f"{sw_name} r{sw_rel}")
parser.add_argument("-v", "--version", help="Lấy số hiệu phiên bản.", action="store_true")
#parser.add_argument("-l", "--launcher", help="Chỉ ra vị trí launcher để cập nhật.", nargs='?', default=[])
args = parser.parse_args()

# cho phép gọi với tham số --version để lấy phiên bản
if args.version:
	print(sw_rel)
	sys.exit(0)

no_login = False # KHÔNG ENABLE TRÊN BẢN PRODUCTION!
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"): no_login = False

import os
import re
import csv
import math
import requests
from requests.exceptions import Timeout
import time
from lxml.html import fromstring
from urllib.parse import unquote
import traceback
import json
import stat
import shutil
import platform
import zipfile
import tarfile
import pyperclip
import functools
import subprocess
import urllib3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from threading import current_thread
import pathlib
try: from subprocess import CREATE_NO_WINDOW
except: pass
import random

# mã hoá AES-256
from base64 import b64encode, b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from hashlib import sha256

# lấy thời gian chạy
from datetime import datetime, timezone
import dateparser
from dateutil.relativedelta import relativedelta

# GUI
from PyQt5.QtWidgets import *
from PyQt5 import QtCore

# selenium
import selenium
se_version = int(selenium.__version__.split('.')[0])

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.firefox.options import Options as FFOptions
from selenium.webdriver.firefox.service import Service as FFService

qt_app = None
qt_win = None

py_legacy = True if sys.version_info[1] <= 7 else False
if py_legacy:
	urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ho tro cac phien ban Python 3 truoc 3.9 (bang cach viet lai PEP 616)
def removeprefix(s, prefix):
	if s.startswith(prefix): return s[len(prefix):]
	else: return s[:]
def removesuffix(s, suffix):
	if suffix and s.endswith(suffix): return s[:-len(suffix)]
	else: return s[:]

# cấu hình (gồm tài khoản, cookies, v.v.)
config = {}
if os.path.isfile("config.json"):
	with open("config.json", "r") as file:
		try: config = json.load(file)
		except: pass

uid_dict = {} # cache UID
cookies = {} # để load trong các request với FB
cookies_aux = {} # dùng để check comment
rq_headers = {} # cho requests

ua_list = []

driver = None # session Chrome

# thông tin tài khoản (để check mỗi lần chạy)
acc_rank = 4405
acc_last_renewed = 0
acc_id = ""

acc_pol_wlist = []
acc_pol_blist = []

if platform.system() == "Windows" and (platform.release() == "XP" or platform.release() == "Vista"):
	os_combo = "windows_lgc-x86"
else:
	os_combo = platform.system().lower() + '-' + platform.machine().lower()

# hàm lấy hạn sử dụng
def get_expire():
	global acc_last_renewed
	expire = datetime.utcfromtimestamp(acc_last_renewed)
	if acc_rank == 0: expire += relativedelta(days=7) # rank 0: dùng thử 7 ngày
	if acc_rank == 1: expire += relativedelta(months=1) # rank 1: bản quyền 1 tháng
	if acc_rank == 2: expire += relativedelta(months=6) # rank 2: bản quyền 6 tháng
	if acc_rank == 3: expire += relativedelta(years=1) # rank 3: bản quyền 1 năm
	return expire

# hàm lấy tên rank
def get_rank():
	if acc_rank == 0: return "dùng thử 7 ngày"
	if acc_rank == 1: return "bản quyền 1 tháng"
	if acc_rank == 2: return "bản quyền 6 tháng"
	if acc_rank == 3: return "bản quyền 1 năm"
	if acc_rank == 255: return "bản quyền vĩnh viễn"
	if acc_rank == 4405: return "developer"

# hàm check hạn sử dụng phần mềm (chạy trước khi check)
def check_acc(app):
	if no_login: return
	global acc_id
	resp = requests.post("https://itsmevjnk.mooo.com/hrng/check.php", data={"id":acc_id}, verify=not py_legacy).text
	if resp == "ERR":
		app.msgbox("critical", "Lỗi kiểm tra bản quyền", "Đã có lỗi trong quá trình xác thực bản quyển.\nHãy thử lại trong ít phút hoặc liên hệ itsmevjnk.work@gmail.com nếu vấn đề chưa được khắc phục.")
		app.exit_sig.emit()
	if resp == "FALSE":
		app.msgbox("critical", "Hết hạn bản quyền", "Bản quyền đã hết hạn.\nVui lòng gia hạn hoặc liên hệ itsmevjnk.work@gmail.com.")
		app.exit_sig.emit()

# thư mục lưu các file kết quả
save_dir = os.path.expanduser("~")
if os.path.isdir(os.path.join(save_dir, "Desktop")): save_dir = os.path.join(save_dir, "Desktop")

relogin = False # chuyển sang True nếu đăng nhập lại

# hàm load trang bằng trình duyệt (để hạn chế load 1 trang nhiều lần)
driver_url = ""
def driver_get(url, force = False):
	global driver_url, driver
	if force or url != driver_url:
		# print("driver_get(): Loading", url)
		driver.get(url)
		driver_url = url

def load_cookies(host, cks):
	global driver
	driver_get(host, force = True)
	# print("Deleting cookies.")
	driver.delete_all_cookies()
	for cookie in cks:
		driver.add_cookie({"name": cookie, "value": cks[cookie]})
	# print("Refreshing.")
	driver.refresh()

# hàm phân tích HTML và tìm element theo xpath
def find_elem(data, xpath):
	return fromstring(data).xpath(xpath)

class login_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(login_worker, self).__init__()
		self.app = app

	def run(self):
		global relogin
		if self.app.lne_email.text() == "" or self.app.lne_password.text() == "" or self.app.lne_cookies.text() == "" or (self.app.chk_cookies_aux.isChecked() and self.app.lne_cookies_aux.text() == ""):
			self.app.msgbox("warning", "Thông báo", "Vui lòng nhập đầy đủ thông tin đăng nhập.")
			self.finished.emit(); return
		self.app.switch_wid_sig.emit(1)
		while self.app.current_wid != 1: time.sleep(0.05)
		self.app.init_stat_sig.emit(0, " ")

		# lưu thông tin đăng nhập
		config["email"] = self.app.lne_email.text()
		config["cookies"] = self.app.lne_cookies.text().replace(" ", "")
		config["cookies_aux"] = self.app.lne_cookies_aux.text().replace(" ", "")

		# load cookies
		global cookies, cookies_aux
		for pair in config["cookies"].split(";"):
			pair_list = pair.split('=')
			if len(pair_list) != 2: continue
			cookies[pair_list[0]] = pair_list[1]
			if pair_list[0] == "_uafec":
				user_agent_str = unquote(pair_list[1]) # lấy user agent
		for pair in config["cookies_aux"].split(";"):
			pair_list = pair.split('=')
			if len(pair_list) != 2: continue
			cookies_aux[pair_list[0]] = pair_list[1]
		
		
		# đăng nhập tài khoản (CHỖ CODE NÀY CẤM LEAK)
		if not no_login:
			cipher = AES.new(sha256("VFWa3c5wg8aRIiui9zI8arSj2mWkni1NQqbpGtEFaf0".encode("ascii")).digest()[:32], AES.MODE_CBC) # key từ app
			enc_data = cipher.encrypt(pad(json.dumps({"email": self.app.lne_email.text(), "pw": self.app.lne_password.text()}).encode("utf-8"), AES.block_size))
			result = b64encode(enc_data + cipher.iv)
			resp_enc = b64decode(requests.post("https://itsmevjnk.mooo.com/hrng/auth.php", data = {"payload": result}, verify=not py_legacy).text)
			cipher = AES.new(sha256("CjIyOuyg_H6turuMVhdR4KkvMbJU5gAxb8Wgyklc_E0".encode("ascii")).digest()[:32], AES.MODE_CBC, resp_enc[-16:]) # key từ server
			resp = json.loads(unpad(cipher.decrypt(resp_enc[:-16]), AES.block_size).decode("utf-8"))
			global acc_rank, acc_last_renewed, acc_id, acc_pol_wlist, acc_pol_blist
			acc_rank = int(resp.get("rank", 0))
			acc_last_renewed = int(resp.get("last_renewed", 0))
			acc_id = resp.get("id", "")
			acc_pol_wlist = [x for x in resp.get("pol_id_wlist", "").split(",") if x]
			acc_pol_blist = [x for x in resp.get("pol_id_blist", "").split(",") if x]
			errno = int(resp.get("errno", -999))
			if errno < 0:
				# if errno != -3 and errno != -4: raise
				self.app.msgbox("critical", "Lỗi đăng nhập", "Sai email hoặc mật khẩu.")
				self.app.switch_wid_sig.emit(0)
				while self.app.current_wid != 0: time.sleep(0.05)
				self.finished.emit(); return
			if abs(datetime.utcnow().replace(tzinfo=timezone.utc).timestamp() - int(resp.get("timestamp", 0))) >= 300: # 5-minute
				self.app.msgbox("critical", "Lỗi đăng nhập", "Thông tin đăng nhập đã hết hạn. Vui lòng thử đăng nhập lại.")
				self.app.switch_wid_sig.emit(0)
				while self.app.current_wid != 0: time.sleep(0.05)
				self.finished.emit(); return
			check_acc(self.app)
			self.app.lbl_license.setText("Email đăng ký: {} ({})".format(self.app.lne_email.text(), get_rank()))
			if acc_rank != 4405 and acc_rank != 255:
				self.app.lbl_license.setText("{}\nHạn sử dụng: {}".format(self.app.lbl_license.text(), str(get_expire())))
		
		# nạp danh sách user agent
		global ua_list
		ua_list.extend(json.loads(requests.get("https://itsmevjnk.mooo.com/hrng/ua.json", verify=not py_legacy).text))

		# hàm tải file lớn
		def download(url, perc, output = "temp_dl"):
			global qt_win
			last_percent = 0
			# print("Downloading", url)
			with open(output, "wb") as f:
				resp = requests.get(url, stream = True, verify=not py_legacy)
				total_len = resp.headers.get("content-length")
				if total_len is None: f.write(resp.content)
				else:
					dl = 0
					total_len = int(total_len)
					init_perc = self.app.pbr_init.value()
					for data in resp.iter_content(chunk_size = total_len // perc):
						dl += len(data)
						f.write(data)
						percent = int((dl / total_len) * perc)
						if percent > last_percent:
							last_percent = percent
							self.app.init_stat_sig.emit(percent + init_perc, " ")
		
		# hàm tải và giải nén file zip
		def download_zip(url, perc, dest = ""):
			download(url, perc)
			with zipfile.ZipFile("temp_dl", "r") as zip:
				zip.extractall(dest)
			os.remove("temp_dl")

		use_firefox = self.app.chk_firefox.isChecked() # sử dụng Firefox thay cho Chromium, TODO: tự động chuyển sang Firefox nếu có vấn đề

		if not relogin: # không khởi động lại trình duyệt nếu đang đăng nhập lại
			global driver
			if use_firefox:
				# hàm so sánh phiên bản theo format maj.min.rev
				def ver_lt(a, b):
					a = a.split('.')
					b = b.split('.')
					if int(a[0]) < int(b[0]):
						return True
					if int(a[0]) == int(b[0]) and int(a[1]) < int(b[1]):
						return True
					if int(a[0]) == int(b[0]) and int(a[1]) == int(b[1]) and int(a[2]) < int(b[2]):
						return True
					return False

				# nền tảng Firefox để tải về
				firefox_os = {
					"windows-amd64": "win64",
					"windows-x86": "win",
					"windows_lgc-x86": "win",
					"linux-x86_64": "linux64",
					"darwin-x86_64": "osx"
				}
				# nền tảng GeckoDriver để tải về
				gkdrv_os = {
					"windows-amd64": "win64",
					"windows-x86": "win32",
					"windows_lgc-x86": "win32",
					"linux-x86_64": "linux64",
					"darwin-x86_64": "macos"
				}
				# lấy danh sách phiên bản GeckoDriver từ docs của Firefox
				gkdrv_ver = []
				for version in find_elem(requests.get("https://firefox-source-docs.mozilla.org/testing/geckodriver/Support.html").text, "//tr[not(ancestor::thead)]"):
					itm = {"version": version[0].text.split()[0], "min": int(version[2].text.split()[0]), "max": version[3].text.split()[0]}
					if itm["max"] == "n/a": itm["max"] = 1000 # số hiệu phiên bản rất lớn
					else: itm["max"] = int(itm["max"])
					gkdrv_ver.append(itm)
				# tên file Firefox
				firefox_exec = {
					"Windows": os.path.join(os_combo, "core", "firefox.exe"),
					"Linux": os.path.join(os_combo, "firefox", "firefox")
				}
				firefox_exec = firefox_exec[platform.system()]
				# tên file GeckoDriver
				gkdrv_exec = os.path.join(os_combo, "geckodriver")
				if platform.system() == "Windows": gkdrv_exec += ".exe"
				# lấy phiên bản Firefox ESR mới nhất (thực ra đoạn này có thể tích hợp với tải file nhưng request sẽ khá lâu)
				remote_ver = find_elem(requests.get("https://ftp.mozilla.org/pub/firefox/releases/").text, '//a[contains(@href, "esr") and not(contains(@href, "-"))]')[-1].text.replace("esr/", "")
				local_ver = config.get("local_ver_ff_" + os_combo, "0.0.0")
				if not os.path.isfile(firefox_exec): local_ver = "0.0.0"
				gkdrv_update = False # chuyển thành True để update GeckoDriver
				local_ver_gd = config.get("local_ver_gd_" + os_combo, "0.0.0")
				if local_ver_gd == "0.0.0": gkdrv_update = True
				if ver_lt(local_ver, remote_ver):
					if local_ver == "0.0.0" or ((time.time() - config.get("last_updated_ff", 0)) >= (7 * 24 * 60 * 60) and self.app.msgbox("question", "Cập nhật Firefox", f"Bạn muốn tải phiên bản Firefox mới nhất ({remote_ver}esr) không?") == QMessageBox.Yes):
						if os.path.isdir(os.path.dirname(firefox_exec)): shutil.rmtree(os.path.dirname(firefox_exec))
						self.app.init_stat_sig.emit(0, f"Đang tải phiên bản Firefox mới nhất ({remote_ver}esr)...")
						download("https://download.mozilla.org/?product=firefox-esr-latest&os={}&lang=en-US".format(firefox_os[os_combo]), 33, "firefox.tmp")
						if platform.system() == "Windows":
							# file cho Windows là file EXE dạng 7-Zip SFX, cần giải nén bằng 7-Zip
							if not os.path.isfile(os.path.join("7z", "7za.exe")): # tải 7za
								if not os.path.isdir("7z"): os.mkdir("7z")
								download_zip("https://www.7-zip.org/a/7za920.zip", 1, "7z")
							self.app.init_stat_sig.emit(33, "Đang giải nén Firefox...")
							subprocess.run([os.path.join("7z", "7za.exe"), "x", "firefox.tmp", "-aoa", f"-o{os_combo}", "core"])
						elif platform.system() == "Linux":
							# file cho Linux là file tar.bz2, cần giải nén bằng thư viện tarfile
							with tarfile.open("firefox.tmp", "r:bz2") as tar:
								tar.extractall(os_combo)
						os.remove("firefox.tmp")
						if platform.system() != "Windows": os.chmod(firefox_exec, os.stat(firefox_exec).st_mode | stat.S_IEXEC)
						config["local_ver_ff_" + os_combo] = remote_ver
						local_ver = remote_ver
						config["last_updated_ff"] = time.time()
						# kiểm tra xem phiên bản GeckoDriver hiện tại còn sử dụng được không
						local_maj = int(local_ver.split('.')[0])
						for version in gkdrv_ver:
							if version["version"] == local_ver_gd:
								if version["min"] > local_maj or version["max"] < local_maj:
									gkdrv_update = True # phiên bản hiện tại không còn sử dụng được, phải upgrade lên bản mới nhất
								break
				gkdrv_update_opt = False # chuyển thành True nếu việc nâng cấp GeckoDriver là không bắt buộc
				remote_ver_gd = "0.0.0"
				local_maj = int(local_ver.split('.')[0])
				# print(local_maj)
				for version in reversed(gkdrv_ver):
					# print(version)
					if version["min"] <= local_maj and version["max"] >= local_maj:
						remote_ver_gd = version["version"]
				if ver_lt(local_ver_gd, remote_ver_gd):
					if not gkdrv_update: gkdrv_update_opt = True
					gkdrv_update = True
				if not os.path.isfile(gkdrv_exec) or gkdrv_update:
					if not gkdrv_update_opt or not os.path.isfile(gkdrv_exec) or self.app.msgbox("question", "Cập nhật GeckoDriver", f"Bạn muốn tải phiên bản GeckoDriver mới nhất ({remote_ver_gd}) không?" == QMessageBox.Yes):
						self.app.init_stat_sig.emit(33, "Đang tải GeckoDriver...")
						gdos = gkdrv_os[os_combo]
						download(f"https://github.com/mozilla/geckodriver/releases/download/v{remote_ver_gd}/geckodriver-v{remote_ver_gd}-{gdos}.zip", 33, "gkdrv.tmp")
						if platform.system() == "Windows":
							# file cho Windows là file zip, cần giải nén bằng thư viện zipfile
							with zipfile.ZipFile("gkdrv.tmp", "r") as zip:
								zip.extractall(os_combo)
						elif platform.system() == "Linux":
							# file cho Linux là file tar.bz2, cần giải nén bằng thư viện tarfile
							with tarfile.open("gkdrv.tmp", "r:bz2") as tar:
								tar.extractall(os_combo)
						os.remove("gkdrv.tmp")
						if platform.system() != "Windows": os.chmod(gkdrv_exec, os.stat(gkdrv_exec).st_mode | stat.S_IEXEC)
						config["local_ver_gd_" + os_combo] = remote_ver_gd
				self.app.init_stat_sig.emit(66, "Đang khởi động Firefox...")
				options = FFOptions()
				options.binary_location = firefox_exec
				try: options.set_preference("general.useragent.override", user_agent_str)
				except NameError: pass
				options.set_preference('browser.migration.version', 9001)
				options.set_preference('permissions.default.stylesheet', 2)
				options.set_preference('permissions.default.image', 2)
				service = FFService(executable_path=gkdrv_exec)
				if acc_rank != 4405 or self.app.msgbox("question", "Dành cho dev", "Hiển thị giao diện Firefox và GeckoDriver?") == QMessageBox.No:
					options.headless = True
				if acc_rank != 4405:
					try: webdriver.common.service.subprocess.Popen = functools.partial(subprocess.Popen, creationflags=CREATE_NO_WINDOW)
					except: pass
				driver = webdriver.Firefox(service=service, options=options)
			else: # update Chromium, ChromeDriver và khởi động webdriver
				# thư mục trên server tương ứng với các hệ điều hành
				chrome_server_dir = {
					"windows-amd64": "Win_x64",
					"windows-x86": "Win",
					"windows_lgc-x86": "Win",
					"linux-x86_64": "Linux_x64",
					"darwin-x86_64": "Mac"
				}
				# tên file zip chứa Chromium
				chrome_server_browser = {
					"windows-amd64": "chrome-win",
					"windows-x86": "chrome-win",
					"windows_lgc-x86": "chrome-win32",
					"linux-x86_64": "chrome-linux",
					"darwin-x86_64": "chrome-mac"
				}
				# tên file zip chứa ChromeDriver
				chrome_server_driver = {
					"windows-amd64": "chromedriver_win32",
					"windows-x86": "chromedriver_win32",
					"windows_lgc-x86": "chromedriver_win32",
					"linux-x86_64": "chromedriver_linux64",
					"darwin-x86_64": "chromedriver_mac64"
				}
				# tên trong danh sách phiên bản
				chrome_ver_entry = {
					"windows-amd64": "win64",
					"windows-x86": "win",
					"linux-x86_64": "linux64",
					"darwin-x86_64": "mac"
				}
				# tên file Chrome
				chrome_exec = os.path.join(os_combo, chrome_server_browser[os_combo])
				if platform.system() == "Darwin":
					chrome_exec = os.path.join(chrome_exec, "Chromium.app", "Contents", "MacOS", "Chromium")
				else:
					chrome_exec = os.path.join(chrome_exec, "chrome")
					if platform.system() == "Windows": chrome_exec += ".exe"
				# tên file ChromeDriver
				if platform.system() == "Windows" and (platform.release() == "XP" or platform.release() == "Vista"):
					chromedrv_exec = os.path.join(os_combo, "chromedriver.exe")
				else:
					chromedrv_exec = os.path.join(os_combo, chrome_server_driver[os_combo], "chromedriver")
					if platform.system() == "Windows": chromedrv_exec += ".exe"
				releases = requests.get("https://raw.githubusercontent.com/Bugazelle/chromium-all-old-stable-versions/master/chromium.stable.json", verify=not py_legacy).json()[chrome_ver_entry[os_combo]]
				darwin_completed = False
				if platform.system() == "Darwin":
					# ưu tiên sử dụng Chrome đã cài
					chrome_exec = ""
					if os.path.isfile("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"):
						chrome_exec = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
					elif os.path.isfile("/Applications/CocCoc.app/Contents/MacOS/CocCoc"):
						chrome_exec = "/Applications/CocCoc.app/Contents/MacOS/CocCoc"
					if chrome_exec != "":
						local_ver = subprocess.check_output([chrome_exec, '--version']).decode('utf8').rstrip().split()[-1].split('.')[0]
						for remote_ver in releases.keys():
							if(remote_ver.startswith(local_ver)):
								remote_ver = releases[remote_ver]["download_position"]
								local_ver = config.get("local_ver_" + os_combo, 0)
								if local_ver != remote_ver:
									self.app.init_stat_sig.emit(33, "Đang tải ChromeDriver...")
									download_zip("https://www.googleapis.com/download/storage/v1/b/chromium-browser-snapshots/o/{}%2F{}%2F{}.zip?alt=media".format(chrome_server_dir[os_combo], remote_ver, chrome_server_driver[os_combo]), 33, os_combo)
									os.chmod(chromedrv_exec, os.stat(chromedrv_exec).st_mode | stat.S_IEXEC)
									config["local_ver_" + os_combo] = remote_ver
								darwin_completed = True
								break
				if not darwin_completed:
					if platform.system() == "Windows" and (platform.release() == "XP" or platform.release() == "Vista"): remote_ver = 369909
					else: remote_ver = releases[next(iter(releases))]["download_position"]
					local_ver = config.get("local_ver_" + os_combo, 0)
					if not os.path.isfile(chrome_exec): local_ver = 0
					if local_ver < remote_ver:
						if local_ver == 0 or ((time.time() - config.get("last_updated", 0)) >= (7 * 24 * 60 * 60) and self.app.msgbox("question", "Cập nhật Chrome", "Bạn muốn tải phiên bản Chrome mới nhất ({}) không?".format(remote_ver)) == QMessageBox.Yes):
							if os.path.isdir(os.path.join(os_combo, chrome_server_browser[os_combo])): shutil.rmtree(os.path.join(os_combo, chrome_server_browser[os_combo]))
							self.app.init_stat_sig.emit(0, "Đang tải phiên bản Chrome mới nhất ({})...".format(remote_ver))
							download_zip("https://www.googleapis.com/download/storage/v1/b/chromium-browser-snapshots/o/{}%2F{}%2F{}.zip?alt=media".format(chrome_server_dir[os_combo], remote_ver, chrome_server_browser[os_combo]), 33, os_combo)
							if not platform.system() == "Windows": os.chmod(chrome_exec, os.stat(chrome_exec).st_mode | stat.S_IEXEC)
							if os.path.isdir(os.path.join(os_combo, chrome_server_driver[os_combo])): shutil.rmtree(os.path.join(os_combo, chrome_server_driver[os_combo])) # xoá ChromeDriver hiện tại
							elif os.path.isfile(chromedrv_exec): os.remove(chromedrv_exec)
							config["local_ver_" + os_combo] = remote_ver
							local_ver = remote_ver
							config["last_updated"] = time.time()
					if not os.path.isfile(chromedrv_exec):
						self.app.init_stat_sig.emit(33, "Đang tải ChromeDriver...")
						if os_combo == "windows_lgc-x86":
							download_zip("https://chromedriver.storage.googleapis.com/2.19/chromedriver_win32.zip", 33, os_combo)
						else:
							download_zip("https://www.googleapis.com/download/storage/v1/b/chromium-browser-snapshots/o/{}%2F{}%2F{}.zip?alt=media".format(chrome_server_dir[os_combo], remote_ver, chrome_server_driver[os_combo]), 33, os_combo)
							if not platform.system() == "Windows": os.chmod(chromedrv_exec, os.stat(chromedrv_exec).st_mode | stat.S_IEXEC)
				self.app.init_stat_sig.emit(66, "Đang khởi động Chrome...")
				options = Options()
				options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
				options.binary_location = chrome_exec
				try: options.add_argument("--user-agent=\"{}\"".format(user_agent_str)) # tôn hành giả, giả hành tôn
				except NameError: pass
				# options.add_argument("--no-sandbox")
				service = Service(chromedrv_exec)
				if acc_rank != 4405 or self.app.msgbox("question", "Dành cho dev", "Hiển thị giao diện Chrome và ChromeDriver?") == QMessageBox.No:
					options.add_argument("--headless")
					options.add_argument("--ignore-certificate-errors") # cho HĐH cũ/chưa update
					options.add_argument("--disable-gpu") # fix lỗi trên Windows, thực ra không cần lắm trên Linux nhưng mà để vào cũng không sao
				if acc_rank != 4405:
					try: webdriver.common.service.subprocess.Popen = functools.partial(subprocess.Popen, creationflags=CREATE_NO_WINDOW)
					except: pass
				driver = webdriver.Chrome(options = options, service = service)

		# load cookies cho Chromium
		self.app.init_stat_sig.emit(83, " ")
		# đăng nhập cookies tài khoản phụ (nếu có)
		if self.app.chk_cookies_aux.isChecked():
			load_cookies("https://m.facebook.com", cookies_aux)
			if len(driver.find_elements(By.XPATH, '//button[@name="login"]')) > 0:
				if self.app.msgbox("question", "Lỗi đăng nhập", "Đăng nhập Facebook bằng tài khoản check comment và tag thất bại. Bạn muốn nhập lại cookies không? (Nếu không, phần mềm sẽ không sử dụng tài khoản này.)") == QMessageBox.Yes:
					self.app.switch_wid_sig.emit(0)
					while self.app.current_wid != 0: time.sleep(0.05)
					relogin = True
					self.finished.emit(); return
				else:
					self.app.chk_cookies_aux.setChecked(False)
		# đăng nhập cookies tài khoản chính
		load_cookies("https://m.facebook.com", cookies)
		if len(driver.find_elements(By.XPATH, '//button[@name="login"]')) > 0:
			if self.app.chk_cookies_aux.isChecked():
				if self.app.msgbox("question", "Lỗi đăng nhập", "Đăng nhập Facebook bằng tài khoản chính thất bại. Bạn muốn nhập lại cookies không? (Nếu không, phần mềm sẽ sử dụng tài khoản check comment và tag thay cho tài khoản này.)") == QMessageBox.No:
					self.app.chk_cookies_aux.setChecked(False)
					cookies = cookies_aux
					load_cookies("https://m.facebook.com", cookies)
					if len(driver.find_elements(By.XPATH, '//button[@name="login"]')) > 0:
						self.app.msgbox("warning", "Lỗi đăng nhập", "Đăng nhập Facebook thất bại. Vui lòng nhập lại cookies mới và thử đăng nhập lại.")
						self.app.switch_wid_sig.emit(0)
						while self.app.current_wid != 0: time.sleep(0.05)
						relogin = True
						self.finished.emit(); return
				else:
					self.app.switch_wid_sig.emit(0)
					while self.app.current_wid != 0: time.sleep(0.05)
					relogin = True
					self.finished.emit(); return
			else:
				self.app.msgbox("warning", "Lỗi đăng nhập", "Đăng nhập Facebook thất bại. Vui lòng nhập lại cookies mới và thử đăng nhập lại.")
				self.app.switch_wid_sig.emit(0)
				while self.app.current_wid != 0: time.sleep(0.05)
				relogin = True
				self.finished.emit(); return
		load_cookies("https://mbasic.facebook.com", cookies)
		
		global rq_headers
		# lưu header cho requests
		self.app.init_stat_sig.emit(99, " ")
		driver_get("https://itsmevjnk.mooo.com/hrng/headers.php")
		rq_headers = json.loads(driver.find_element(By.XPATH, '//body').text)
		rq_headers.pop("Connection", None)
		rq_headers.pop("Host", None)
		rq_headers.pop("Cookie", None)
		rq_headers.pop("Sec-Fetch-Site", None)
		rq_headers.pop("Sec-Fetch-Mode", None)
		rq_headers.pop("Sec-Fetch-User", None)
		rq_headers.pop("Sec-Fetch-Dest", None)
		rq_headers.pop("Content-Length", None)
		rq_headers.pop("Content-Type", None)
		rq_headers['Accept-Encoding'] = 'gzip, deflate'
		# print(rq_headers)

		# lưu cấu hình
		self.app.init_stat_sig.emit(100, "Đang lưu cấu hình.")
		with open("config.json", "w") as file: json.dump(config, file)

		lbl_version_text = "Python {} @ {}\nSelenium {}, ".format(platform.python_version(), platform.platform(), selenium.__version__)
		if use_firefox:
			lbl_version_text += "Mozilla Firefox {}esr, GeckoDriver v{}".format(config["local_ver_ff_" + os_combo], config["local_ver_gd_" + os_combo])
		else:
			lbl_version_text += "Chromium build {}".format(config["local_ver_" + os_combo])
		self.app.lbl_version.setText(lbl_version_text)

		# driver.maximize_window()

		# chuyển vào giao diện chính và kết thúc thread
		self.app.switch_wid_sig.emit(2)
		while self.app.current_wid != 2: time.sleep(0.05)
		self.finished.emit()

# hàm tải 1 trang Facebook
def fb_load(url):
	if url.startswith('/'): url = "https://mbasic.facebook.com" + url
	return removeprefix(requests.get(url, cookies=cookies, headers=rq_headers, verify=not py_legacy).text, '<?xml version="1.0" encoding="utf-8"?>')

# hàm lấy UID (từ link -> từ cache -> không đăng nhập -> có đăng nhập)
def get_uid_preprocess(link):
	link = removesuffix(link, '/')
	link = "".join(link.split())
	link = re.sub("home\.php.*\/", "", link)
	link = re.sub("\/\?.*$", "", link)
	link = re.sub("(^.*\.((co.*)|(me))\/)|(&.*$)", "", link)
	link = removeprefix(link, '/')
	if not link.startswith("profile.php"): link = re.sub("\?.*$", "", link)
	return link

def get_uid(link, app = None):
	link = get_uid_preprocess(link)
	if link.count("/") != 0: return "" # link không hợp lệ
	global uid_dict
	uid = None
	# thử lấy UID trực tiếp từ link
	if link.startswith("profile.php"):
		#self.app.bump_stat_sig.emit(link + " có sẵn UID")
		uid = re.sub('^.*\?id=', "", link)
		if uid == "profile.php": return ""
		return str(uid)
	# thử tìm UID trong cache
	link = re.sub("(^.*\/)|(\?.*$)", "", link)
	if link in uid_dict:
		#self.app.bump_stat_sig.emit(link + " có sẵn trong cache")
		return uid_dict[link]
	# lấy UID qua dịch vụ lấy UID (có thể lâu nhưng không sợ bị hạn chế)
	while True:
		# hàm lấy UID bằng dịch vụ
		got_uid = 0
		def uid_serv(link):
			global rq_headers
			def attempt(url, data, xpath):
				global rq_headers
				for i in range(3): # thử tối đa 3 lần để tránh bị ban IP
					rq_headers["User-Agent"] = random.choice(ua_list)
					try: page = fromstring(requests.post(url, data = data, headers = rq_headers, timeout = 10, verify=not py_legacy).text)
					except:
						return []
					if len(page.xpath('//a[@href="https://www.cloudflare.com/5xx-error-landing"]')) == 0:
						try: got_uid = len(page.xpath(xpath))
						except: got_uid = 0
						return page.xpath(xpath)
			link = "https://facebook.com/" + link
			orig_ua = rq_headers["User-Agent"]
			uid_list = attempt("https://id.atpsoftware.vn/", {"linkCheckUid": link}, "//div[@id='menu1']/textarea/text()")
			if len(uid_list) == 0: uid_list = attempt("https://findidfb.com/#", {"url": link}, "//div[@class='alert alert-success alert-dismissable']/b/text()")
			if len(uid_list) == 0: uid_list = attempt("https://lookup-id.com/#", {"fburl": link, "check": "Lookup"}, "//span[@id='code']/text()")
			rq_headers["User-Agent"] = orig_ua
			return uid_list
		uid_list = uid_serv(link)
		if len(uid_list) != 0 and uid_list[0].isdigit():
			uid = str(uid_list[0])
			break
		else:
			# cách cuối (và cũng là cách nguy hiểm nhất): mở link để check.
			page = fb_load('/' + link)
			for block_link in find_elem(page, '//a[starts-with(@href, "/privacy/touch/block/confirm/?bid=")]/@href'):
				uid = str(re.sub('(^.*\?bid=)|(&.*$)', "", block_link))
				uid_dict[link] = uid
				return uid
			# tìm UID cho page
			for more_link in find_elem(page, '//a[starts-with(@href, "/pages/more/")]/@href'):
				uid = str(re.sub('\/.*$', "", re.sub('^\/pages\/more\/', "", more_link)))
				uid_dict[link] = uid
				return uid
			# check xem có bị Facebook ratelimit không
			if len(find_elem(page, '//a[@href="https://www.facebook.com/help/177066345680802"]')) > 0:
				if app != None: app.msgbox("critical", "Lỗi", "Tài khoản đang dùng đã bị Facebook chặn mở link. Vui lòng đăng nhập lại bằng tài khoản khác không bị chặn.\nPhần mềm sẽ thoát.")
				global config
				config["cookies"] = ""
				with open("config.json", "w") as file: json.dump(config, file)
				if app != None: app.do_exit()
			if app == None or app.msgbox("question", "Không lấy được UID", "Không thể lấy được UID của link https://facebook.com/{}.\nThử lại?".format(link)) == QMessageBox.No: return ""
	# print(link + " -> " + uid)
	uid_dict[link] = uid
	return uid

def check_ratelimit(app):
	global driver
	if len(driver.find_elements(By.XPATH, '//a[@href="https://www.facebook.com/help/177066345680802"]')) > 0:
		app.msgbox("critical", "Lỗi", "Tài khoản đang dùng đã bị Facebook chặn mở link. Vui lòng đăng nhập lại bằng tài khoản khác không bị chặn.\nPhần mềm sẽ thoát.")
		global config
		config["cookies"] = ""
		with open("config.json", "w") as file: json.dump(config, file)
		app.do_exit()

# các hàm dùng cho check bump và confirm
def get_ids(url, app):
	post_id = ""
	author_uid = ""
	page_id = ""
	driver_get(url)
	check_ratelimit(app)
	if driver.current_url.startswith("https://m.facebook.com/watch/"):
		# video trên watch, cần lấy link bài đăng chứa video
		# with open("watch.html", "wb") as file: file.write(bytes(post.text, encoding="utf-8"))
		try:
			post_link = driver.find_element(By.XPATH, '//div[@id="mobile_injected_video_feed_pagelet"]//a[starts-with(@href, "/story.php")]')
			driver_get(post_link.get_attribute("href"))
		except NoSuchElementException:
			pass
	# with open('post.html', 'wb') as f: f.write(bytes(post.text, encoding='utf-8'))
	try:
		post_id = removeprefix(driver.find_element(By.XPATH, '//div[starts-with(@id, "ufi_")]').get_attribute("id"), "ufi_")
	except NoSuchElementException:
		pass
	try:
		author_uid = removeprefix(driver.find_element(By.XPATH, '//div[starts-with(@data-sigil, "feed_story_ring")]').get_attribute("data-sigil"), "feed_story_ring")
	except NoSuchElementException:
		pass
	try:
		page_id = re.sub("(^.*\?groupid=)|(&.*$)", "", driver.find_element(By.XPATH, '//a[contains(@href, "groupid")]').get_attribute("href"))
	except NoSuchElementException:
		page_id = author_uid
	return (page_id, post_id, author_uid)


# hàm spam click element đến khi element không còn tồn tại
def spam_click(xpath, app):
	global driver
	try: see_next_elem = driver.find_element(By.XPATH, xpath)
	except NoSuchElementException: return
	while True:
		try:
			driver.execute_script("arguments[0].click()", see_next_elem)
			see_next_elem = driver.find_element(By.XPATH, xpath)
		except StaleElementReferenceException:
			try: see_next_elem = driver.find_element(By.XPATH, xpath)
			except NoSuchElementException: break
			continue
		except NoSuchElementException:
			break
	time.sleep(0.5) # chờ load hẳn
	check_ratelimit(app)

poll_name = {} # dict chứa tên các lựa chọn trong poll
def check_poll(members, id, sig = None, app = None, pbr_val = 0):
	global poll_name, driver
	poll_name = {}
	driver_get("https://m.facebook.com/" + id)
	check_ratelimit(app)
	if sig != None:
		cnt = pcnt = cnt_total = 0
		og_perc = ls_perc = pbr_val
	spam_click('//a[starts-with(@href, "/a/questions/see_more_options.php")]', app)
	for row in driver.find_elements(By.XPATH, '//div[@class="mPollRow"]'):
		try: name = row.find_element(By.XPATH, './/span[@class="mfsm fcb"]').text
		except: name = "[không lấy được tên]"
		poll_name[row.get_attribute("id")] = name
	for id in poll_name:
		driver_get("https://m.facebook.com/browse/questions/option/voters/?oid=" + str(id))
		spam_click('//a[starts-with(@href, "/browse/questions/option/voters/")]', app)
		elem = driver.find_elements(By.XPATH, '//a[@class="_4kk6 _5b6s"]')
		for user in elem:
			uid = get_uid(user.get_attribute("href"))
			if uid == "0": continue
			if sig != None:
				cnt += 1
				cnt_total += 1
				out_perc = -1
				perc = round(og_perc + (pcnt + cnt / len(elem)) * 33 / len(poll_name))
				if ls_perc < perc:
					ls_perc = out_perc = perc
				sig.emit(out_perc, "Đang check poll... (đã quét được {} lựa chọn, tổng {} TK)".format(pcnt, cnt_total))
			for mem in members:
				if uid in mem["UID"]:
					mem["Poll"] = mem.get("Poll", 0) + 1
					mem["poll_" + str(id)] = 1
			if sig != None:
				cnt = 0
				pcnt += 1
	for mem in members:
		mem["Poll"] = mem.get("Poll", 0)
		for id in poll_name: mem["poll_" + str(id)] = mem.get("poll_" + str(id), 0)

total_cmt = mem_cmt = 0
def check_cmt(members, id, chk_reply = False, chk_mention = False, mention_exc = False, use_aux = False, sig = None, app = None):
	global driver, cookies, cookies_aux
	if use_aux: load_cookies("https://m.facebook.com/" + id, cookies_aux)
	check_ratelimit(app)
	if sig != None:
		cnt = 0
	global total_cmt, mem_cmt
	total_cmt = mem_cmt = 0
	reverse_dir = False # see_prev thay vì see_next
	for mem in members:
		mem["cmt"] = {}
		if chk_mention: mem["cmt_mentions"] = {}
		mem["repl_id"] = []
	n = 0
	last_cnt = 0
	while True:
		try:
			if reverse_dir: see_next_elem = driver.find_element(By.XPATH, '//*[starts-with(@id,"see_prev")]/a')
			else: see_next_elem = driver.find_element(By.XPATH, '//*[starts-with(@id,"see_next")]/a')
		except NoSuchElementException:
			if n == 0 and not reverse_dir:
				reverse_dir = True
				try: see_next_elem = driver.find_element(By.XPATH, '//*[starts-with(@id,"see_prev")]/a')
				except NoSuchElementException: break
			else: break
		driver.execute_script("arguments[0].click()", see_next_elem)
		time.sleep(0.5)
		while True:
			try:
				if reverse_dir: driver.find_element(By.XPATH, '//*[starts-with(@id,"see_prev") and contains(@class, "saving")]')
				else: driver.find_element(By.XPATH, '//*[starts-with(@id,"see_next") and contains(@class, "saving")]')
			except NoSuchElementException: break
		try:
			cnt = len(driver.find_elements(By.XPATH, '//div[contains(@class, "_2a_j")]'))
		except NoSuchElementException:
			cnt = 0
		if cnt == last_cnt:
			print("done", cnt)
			break
		else:
			last_cnt = cnt
			print("next")
        
	while True:
		# cách dưới chỉ làm được với comment
		# hiển thị mọi bình luận (kể cả reply) do Facebook hiển thị lặp reply như comment
		spam_click('//a[starts-with(@href, "/comment/replies/")]', app)
		for user in driver.find_elements(By.XPATH, '//div[contains(@class, "_2a_j")]'):
			try: uid = removeprefix(user.get_attribute("data-sigil"), "feed_story_ring")
			except: continue
			# lưu UID vào cache
			try: link = re.sub("(\?$)|(\&$)", "", re.sub("fref=.*$", "", user.find_element(By.XPATH, ".//div/a").get_attribute("href")))
			except: continue
			if "profile.php" not in link:
				link = re.sub("(^.*\/)|(\?.*$)", "", link)
			# check xem có phải reply không
			parent = user.find_element(By.XPATH, "..")
			total_cmt += 1
			counted = False
			for mem in members:
				if uid in mem["UID"]:
					cmt_body = user.find_element(By.XPATH, '..//div[@data-sigil="comment-body"]')
					cmt_id = json.loads(parent.get_attribute("data-store"))["token"]
					if chk_mention:
						mem["cmt_mentions"][cmt_id] = set()
						try:
							for link in cmt_body.find_elements(By.XPATH, "./a"):
								url = link.get_attribute("href")
								try: url_text = link.text
								except: url_text = ""
								if url.startswith("https://m.facebook.com/") and url_text not in url and get_uid_preprocess(url).count("/") == 0:
									mem["cmt_mentions"][cmt_id].add(get_uid_preprocess(url)) # chỉ lấy tag
						except NoSuchElementException: pass
					try: content = cmt_body.text
					except: content = "[không lấy được nội dung]"
					if "inline-reply" in parent.get_attribute("data-sigil"):
						if not cmt_id in mem["repl_id"]: mem["repl_id"].append(cmt_id)
					mem["cmt"][cmt_id] = content
					if not counted:
						counted = True
						mem_cmt += 1
				global uid_dict
				if link not in uid_dict: uid_dict[link] = uid
			if sig != None:
				cnt += 1
				sig.emit(-1, "Đang check comment... (đã quét được {} cmt)".format(cnt))
				qt_app.processEvents()
		break
	# with open("members.txt", "w", encoding="utf-8-sig") as f: f.write(str(members))
	for mem in members:
		uid = str(mem["UID"])
		# xoá các reply
		if not chk_reply:
			for cmt_id in mem["repl_id"]:
				mem["cmt"].pop(cmt_id)
				if chk_mention: mem["cmt_mentions"].pop(cmt_id)
		for cmt_id in mem["cmt"]:
			mem["cmt_" + cmt_id] = mem["cmt"][cmt_id]
		mem["Comment"] = len(mem["cmt"])
		# xử lý tag
		if chk_mention:
			uid_diff = []
			if mention_exc:
				for m in members:
					uid_diff.extend(m["UID"])
			mem["Tag"] = len(set().union(*[mem["cmt_mentions"][id] for id in mem["cmt_mentions"]]) - set(uid_diff))
			# print(mem["UID"], mem["Tag"], [mem["cmt_mentions"][id] for id in mem["cmt_mentions"]])
			del mem["cmt_mentions"]
		del mem["cmt"]
		del mem["repl_id"]
		#print(mem)
	if use_aux: load_cookies("https://m.facebook.com", cookies)

total_react = mem_react = 0
def check_react(members, id, author, sig = None, pbr_val = 0, app = None):
	global driver
	driver_get("https://m.facebook.com/ufi/reaction/profile/browser/?ft_ent_identifier=" + id)
	check_ratelimit(app)
	spam_click('//*[@id="reaction_profile_pager"]/a', app)
	global total_react, mem_react
	total_react = mem_react = 0
	if sig != None:
		cnt = 0
		og_perc = ls_perc = pbr_val
	elem = driver.find_elements(By.XPATH, '//div[@data-sigil="undoable-action marea"]')
	total_react = len(elem)
	has_author = 0
	if str(author) in list(v["UID"] for v in members): has_author = 1
	for user in elem:
		# đường tắt để lấy UID
		try: uid = str(json.loads(user.find_element(By.XPATH, ".//*[contains(@data-store, '\"id\":')]").get_attribute("data-store"))["id"])
		except NoSuchElementException:
			try: str(json.loads(user.find_element(By.XPATH, ".//*[contains(@data-store, '\"subject_id\":')]").get_attribute("data-store"))["subject_id"])
			except NoSuchElementException: uid = get_uid(user.find_element(By.XPATH, ".//div[@class='_4mn c']/a").get_attribute("href"), app)
		if uid == "0": continue
		if sig != None:
			cnt += 1
			out_perc = -1
			perc = round(og_perc + (cnt / len(elem)) * 33)
			if ls_perc < perc:
				ls_perc = out_perc = perc
			sig.emit(out_perc, "Đang check react... (đã quét được {} TK)".format(cnt))
		counted = False
		for mem in members:
			if uid in mem["UID"]:
				mem["React"] = 1
				if not counted:
					counted = True
					mem_react += 1
		# tiết kiệm thời gian bằng cách dừng check khi mọi người đều đã react
		total_mem_react = sum(v.get("React", 0) for v in members) + has_author
		if total_mem_react == len(members): break
	for mem in members:
		mem["React"] = mem.get("React", 0)

total_share = mem_share = 0
def check_share(members, id, author, sig = None, pbr_val = 0, app = None):
	global driver
	driver_get("https://m.facebook.com/browse/shares?id=" + id)
	check_ratelimit(app)
	spam_click('//*[@id="m_more_item"]/a', app)
	global total_share, mem_share
	total_share = mem_share = 0
	if sig != None:
		cnt = 0
		og_perc = ls_perc = pbr_val
	elem = driver.find_elements(By.XPATH, '//*[@class="_4mn c"]/a')
	total_share = len(elem)
	has_author = 0
	if str(author) in list(v["UID"] for v in members): has_author = 1
	for mem in members: mem["share_uid"] = set()
	for user in elem:
		uid = get_uid(user.get_attribute("href"), app)
		if uid == "0": continue
		if sig != None:
			cnt += 1
			out_perc = -1
			perc = round(og_perc + (cnt / len(elem)) * 33)
			if ls_perc < perc:
				ls_perc = out_perc = perc
			sig.emit(out_perc, "Đang check share... (đã quét được {} TK)".format(cnt))
		counted = False
		for mem in members:
			if uid in mem["UID"]:
				mem["share_uid"].add(uid)
				if not counted:
					counted = True
					mem_share += 1
		# tiết kiệm thời gian bằng cách dừng check khi mọi người đều đã share
		total_mem_share = sum(v.get("Share", 0) for v in members) + has_author
		if total_mem_share == len(members): break
	for mem in members:
		mem["Share"] = len(mem["share_uid"])

def scroll_elem(xpath):
	global driver
	try: elem = driver.find_element(By.XPATH, xpath)
	except NoSuchElementException: return None # không có element
	while True:
		try: ActionChains(driver).move_to_element(elem).perform()
		except StaleElementReferenceException: return 0 # hoàn thành
		time.sleep(0.1) # chờ load

def get_scap(members, id, sig = None, app = None):
	global driver
	n = 0
	ptime = 0 # mốc thời gian đăng bài
	for mem in members:
		mem["Cap share"] = ""
		if mem.get("Share", 0) != 0:
			# thành viên đã share
			idx = 0 # số chỉ bắt đầu
			for uid in mem["share_uid"]:
				driver_get("https://m.facebook.com/" + uid, force = True)
				check_ratelimit(app)
				while True:
					posts = driver.find_elements(By.XPATH, "//article")
					stop_finding = False
					for post in posts[idx:]:
						# lấy mốc thời gian đăng bài và kiểm tra xem bài có cũ hơn bài gốc không
						stime = dateparser.parse(post.find_element(By.XPATH, './div/header//div[@data-sigil="m-feed-voice-subtitle"]//abbr').text)
						if math.floor(stime.timestamp()) < ptime:
							stop_finding = True
							break
						if f"original_content_id.{id}" in post.get_attribute("data-store"):
							# tìm được bài share
							if ptime == 0:
								# lấy mốc thời gian đăng bài gốc
								data_ft = json.loads(post.get_attribute("data-ft"))["page_insights"]
								ptime = data_ft[next(iter(data_ft.keys()))]["attached_story"]["post_context"]["publish_time"]
							# lấy cap share
							cap = post.find_element(By.XPATH, './div/div[@data-ft=\'{"tn":"*s"}\']').text
							stime_r = str(stime)
							mem["Cap share"] += f"[{stime_r}] {cap}\n"
					if stop_finding: break # dừng ngay vì không tìm được bài share sau khi bài gốc được đăng
					idx = len(posts)
					if scroll_elem("//div[@class='_52jj _2ph_']") == None: break
			mem["Cap share"] = removesuffix(mem["Cap share"], "\n")
			# print(mem["Cap share"])
			n += 1
			if sig != None: sig.emit(-1, f"Đang lấy cap share... (đã quét được {n} thành viên)")
		

def cp34_iter_rows_val(sheet, min_row, max_row = -1):
	if max_row == -1: max_row = sheet.max_row
	for row in range(min_row, max_row + 1):
		cells = (sheet.cell(row = row, column = column) for column in range(1, sheet.max_column + 1))
		yield tuple(cell.value for cell in cells)

# đọc file
def read_file(fname):
	ret = []
	if pathlib.Path(fname).suffix.lower() == ".xlsx":
		# XLSX
		wb = load_workbook(filename = fname)
		sheet = wb.active
		try:
			r_iter = sheet.iter_rows(min_row = 1, max_row = 1, values_only = True)
		except TypeError:
			r_iter = cp34_iter_rows_val(sheet, 1, 1)
		title = list(r_iter)[-1]
		# print(title)
		try:
			r_iter = sheet.iter_rows(min_row = 2, values_only = True)
		except TypeError:
			r_iter = cp34_iter_rows_val(sheet, 1)
		for row in r_iter:
			# print(row)
			rdict = {}
			for idx, itm in enumerate(title):
				rdict[itm] = row[idx]
			ret.append(rdict)
	else:
		# CSV
		with open(fname, "r", newline="", encoding="utf-8-sig") as f:
			ret = list(csv.DictReader(f))
	for row in ret:
		for k in row.keys():
			if row[k] != None: row[k] = removesuffix(removeprefix(str(row[k]), '="'), '"')
		row.pop(None, None)
		row.pop('', '')
	ret = [row for row in ret if len({k: v for k, v in row.items() if v != None and v != ''}) > 0]
	return ret

def xlsx_sanitize(row):
	for k, v in enumerate(row):
		row[k] = ILLEGAL_CHARACTERS_RE.sub(r'', v)
		row[k] = re.sub("^=", " =", v)
		row[k] = re.sub(r'[\x00-\x09\x0b-\x1f\x7f-\x9f]', '', v) # bỏ mọi control character (trừ xuống dòng)
	return row

# ghi file
def write_file(fname, title, col, data, stats = None):
	for row in data:
		for k in row.keys():
			if type(row[k]) == int:
				row[k] = str(row[k])
			if type(row[k]) == list or type(row[k]) == set:
				row[k] = ";".join(row[k])
	if col == None:
		col = {}
		for k in data[0].keys():
			if k == None: continue
			col[k] = k
	# print(data)
	if pathlib.Path(fname).suffix.lower() == ".xlsx":
		# XLSX
		wb = Workbook()
		ws = wb.active
		# ws.title = title
		ws['A1'] = ""
		ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = len(col))
		out = [list(col.keys())]
		for r in data:
			row = []
			for k in col.keys():
				if col[k].endswith('*'):
					# wildcard (tất cả key bắt đầu bằng k)
					for k2 in r.keys():
						if k2.startswith(removesuffix(col[k], '*')):
							row.append(r[k2])
				else:
					row.append(r.get(col[k], ""))
			row = xlsx_sanitize(row)
			out.append(row)
		if type(stats) == list: out.extend(stats)
		# print(out)
		for line in out:
			line = xlsx_sanitize(line)
			# print(line)
			ws.append(line)
		fill = PatternFill(start_color = "FF9900", end_color = "FF9900", fill_type = "solid")
		bold = Font(bold = True)
		center = Alignment(horizontal = "center")
		ws['A1'].fill = fill
		ws['A1'].font = bold
		ws['A1'].alignment = center
		for x in range(len(col)):
			ws.cell(2, x + 1).fill = fill
			ws.cell(2, x + 1).font = bold
			ws.cell(2, x + 1).alignment = center
		for column_cells in ws.columns:
			length = max(len(str(cell.value if type(cell.value) != None else "")) for cell in column_cells)
			ws.column_dimensions["A"].width = length
		ws['A1'] = title
		ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = len(col))
		wb.save(fname)
	else:
		# CSV
		with open(fname, 'w', newline = '', encoding = 'utf-8-sig') as f:
			with csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL) as out:
				out.writerow([title])
				out.writerow(list(col.keys()))
				for r in data:
					row = []
					for k in col.keys():
						if col[k].endswith('*'):
							# wildcard (tất cả key bắt đầu bằng k)
							for k2 in r.keys():
								if k2.startswith(removesuffix(col[k], '*')):
									row.append(r[k2])
						else:
							row.append(r.get(col[k], ""))
					out.writerow(row)
				if type(stats) == list:
					for row in stats:
						out.writerow(row)

# hàm xử lý UID và link trong danh sách thành viên
def preprocess_mlist(member, app = None):
	for mem in member:
		# print(mem)
		uid = []
		if mem.get("UID", None) != None:
			for v in mem.get("UID", "").replace(" ", "").split(";"):
				uid.append(v)
		if mem.get("Link", None) != None:
			for v in mem.get("Link", "").replace(" ", "").split(";"):
				if v: uid.append(get_uid(v, app))
		mem["UID"] = list(dict.fromkeys([str(v) for v in uid if v]))
		#print(mem)
	return member

class getuid_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(getuid_worker, self).__init__()
		self.app = app

	def run(self):
		users_fname = self.app.lne_uid_in.text().replace("\\", "/").replace("\"", "")
		output_fname = self.app.lne_uid_out.text().replace("\\", "/").replace("\"", "")
		if self.app.lne_uid_link.text() == "" and users_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập link Facebook hoặc danh sách link!")
			self.finished.emit(); return
		check_acc(self.app)
		self.app.uid_stat_sig.emit(0, "  ", "  ")
		self.app.wid_state_sig.emit(False)
		tstart = time.time()

		if self.app.lne_uid_link.text() != "":
			self.app.uid_stat_sig.emit(-1, str(get_uid(self.app.lne_uid_link.text())), " ")
		
		if users_fname != "":
			if output_fname != "":
				users = read_file(users_fname)
				if users == None or len(users) == 0:
					self.app.msgbox("critical", "Lỗi", "Danh sách link không hợp lệ!")
				else:
					n0 = 1 if self.app.lne_uid_link.text() != "" else 0
					n = n0
					for user in users:
						self.app.uid_stat_sig.emit(int(n / (n0 + len(users)) * 100), " ", "Đã lấy được UID của {} tài khoản".format(n))
						if user.get("Link", None): user["UID"] = str(get_uid(user["Link"]))
						else: user["UID"] = ""
						n += 1
					write_file(output_fname, "KẾT QUẢ LẤY UID", None, users)
					self.app.lne_uid_out.setText(os.path.join(save_dir, datetime.now().strftime("uid-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
			else:
				self.app.msgbox("warning", "Thông báo", "Chưa nhập file kết quả!")
		global uid_dict
		uid_dict = {}
		
		self.app.uid_stat_sig.emit(100, " ", "Thời gian thực hiện: {} giây".format(time.time() - tstart))

		if users_fname != "" and output_fname != "":
			self.app.msgbox("information", "Thông báo", "Lấy UID hoàn tất. Kết quả cho các link trong file được lưu vào file {} (trong folder {}).".format(os.path.basename(self.app.lne_uid_out.text()), os.path.dirname(self.app.lne_uid_out.text())))
			self.app.lne_uid_out.setText(os.path.join(save_dir, datetime.now().strftime("uid-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
		
		self.app.wid_state_sig.emit(True)
		self.finished.emit()

class mlist_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(mlist_worker, self).__init__()
		self.app = app

	def get_mlist(self, link, uid = False):
		mlist = []
		if "/t/" in link:
			# group chat
			driver_get("https://m.facebook.com/messages/participants/?tid=cid.g." + re.sub("(^.*\/t\/)|(\/.*$)", "", link))
			elem = driver.find_elements(By.XPATH, '//div[@data-sigil="marea"]/a[@data-sigil="touchable"]')
			total_mem = len(elem)
			n = 0
			for mem in elem:
				mdict = {}
				try: mdict["Tên"] = mem.find_element(By.XPATH, './/strong').text
				except: mdict["Tên"] = ""
				mdict["Link"] = mem.get_attribute("href")
				if uid: mdict["UID"] = str(get_uid(mdict["Link"], self.app))
				mlist.append(mdict)
				n += 1
				self.app.mlist_stat_sig.emit(int(n / total_mem * 100), "Đã truy xuất {} thành viên.".format(n))
		elif "/groups/" in link:
			# group FB
			driver_get("https://mbasic.facebook.com/groups/{}?view=members".format(re.sub("(^.*\/groups\/)|(\/.*$)", "", link)))
			try: link = re.sub("\&.*$", "", driver.find_element(By.XPATH, '//a[contains(@href, "&start=0&listType=list_nonfriend_nonadmin")]').get_attribute("href"))
			except: return []
			# print(link)
			driver_get(link)
			self.app.mlist_stat_sig.emit(0, "Đang tải danh sách thành viên.")
			hlast = driver.execute_script("return document.body.scrollHeight")
			tlast = 0
			while True:
				driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
				hnew = driver.execute_script("return document.body.scrollHeight")
				if hnew == hlast:
					if tlast == 0: tlast = time.time()
					if time.time() - tlast >= 2: break
					else: time.sleep(0.2)
				else:
					tlast = 0
					hlast = hnew
			n = 0
			total_mem = 0
			while True:
				elem = driver.find_elements(By.XPATH, '//table[contains(@id, "member_")]')
				total_mem += len(elem)
				# print(total_mem)
				for mem in elem:
					mdict = {}
					if uid: mdict["UID"] = removeprefix(mem.get_attribute("id"), "member_")
					try:
						link_elem = mem.find_element(By.XPATH, './/h3/a')
						mdict["Tên"] = link_elem.text
						mdict["Link"] = link_elem.get_attribute("href")
					except:
						mdict["Tên"] = ""
						mdict["Link"] = ""
					# print(mdict)
					mlist.append(mdict)
					n += 1
					self.app.mlist_stat_sig.emit(int(n / total_mem * 100), "Đã truy xuất {} thành viên.".format(n))
				try: driver_get(driver.find_element(By.XPATH, '//div[@id="m_more_item"]/a').get_attribute("href"))
				except NoSuchElementException: break

		else: return None
		return mlist

	def run(self):
		cmp_fname = self.app.lne_mlist_cmp.text().replace("\\", "/").replace("\"", "")
		out_fname = self.app.lne_mlist_out.text().replace("\\", "/").replace("\"", "")
		if self.app.lne_mlist_link.text() == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập link nhóm!")
			self.finished.emit(); return
		if out_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập tên file kết quả!")
			self.finished.emit(); return
		if self.app.chk_mlist_cmp.isChecked() and cmp_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập tên file thành viên để đối chiếu!")
			self.finished.emit(); return
		check_acc(self.app)
		self.app.wid_state_sig.emit(False)
		tstart = time.time()

		mlist = self.get_mlist(self.app.lne_mlist_link.text(), self.app.chk_mlist_cmp.isChecked() or self.app.chk_mlist_uid.isChecked())
		if mlist == None:
			self.app.msgbox("critical", "Lỗi", "Link nhóm không hợp lệ!")
			self.finished.emit(); return

		if self.app.chk_mlist_cmp.isChecked():
			cmp = preprocess_mlist(read_file(cmp_fname), self.app)
			cmp_uid = []
			for mem in mlist: cmp_uid.append(mem["UID"])
			# print(cmp_uid)
			for mem in cmp:
				uid = mem.get("UID", [])
				if len(uid):
					try: uid = get_uid(mem["Link"], self.app)
					except:
						mem["Có trong group"] = "Không rõ"
						continue
				#print(mem)
				#print(uid)
				mem["Có trong group"] = "Có" if uid in cmp_uid else "Không"
			write_file(out_fname, "DANH SÁCH THÀNH VIÊN", None, cmp)
		else: write_file(out_fname, "DANH SÁCH THÀNH VIÊN", None, mlist)

		global uid_dict
		uid_dict = {}

		self.app.mlist_stat_sig.emit(100, "Thời gian thực hiện: {} giây".format(time.time() - tstart))
		self.app.msgbox("information", "Thông báo", "Xuất danh sách thành viên hoàn tất. Kết quả được lưu vào file {} (trong folder {}).".format(os.path.basename(self.app.lne_mlist_out.text()), os.path.dirname(self.app.lne_mlist_out.text())))
		self.app.lne_mlist_out.setText(os.path.join(save_dir, datetime.now().strftime("mlist-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))

		self.app.wid_state_sig.emit(True)
		self.finished.emit()

class chkbump_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(chkbump_worker, self).__init__()
		self.app = app

	def run(self):
		check_acc(self.app)

		members_fname = self.app.lne_bump_memfile.text().replace("\\", "/").replace("\"", "")
		output_fname = self.app.lne_bump_outfile.text().replace("\\", "/").replace("\"", "")
		cmt_min = self.app.sbx_bump_cmtmin.value()
		mnt_min = self.app.sbx_bump_mntmin.value()
		post_url = re.sub("^.*com\/", "", self.app.lne_bump_post.text())

		# check các trường tên và link
		if members_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập file thành viên!")
			self.finished.emit(); return
		if output_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập file kết quả!")
			self.finished.emit(); return
		if post_url == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập link post!")
			self.finished.emit(); return
		
		# hiển thị cảnh báo nếu lấy cap share
		if self.app.chk_bump_scap.isChecked():
			if self.app.msgbox("question", "Thông báo", "Tính năng lấy cap share cần truy cập vào tài khoản Facebook của mỗi người đã share và có thể dẫn tới tài khoản bạn bị Facebook giới hạn. Vì vậy, chức năng này chỉ nên dùng với các nhóm nhỏ.\nTiếp tục?") == QMessageBox.No:
				self.app.chk_bump_scap.setChecked(False)

		# load file thành viên
		self.app.bump_stat_sig.emit(0, "Đang đọc và xử lý file thành viên...")
		members = preprocess_mlist(read_file(members_fname), self.app)
		title = {}
		for t in members[0].keys():
			if type(t) != str: continue
			title[t] = t

		# hàm xuất trạng thái bump (bit 0: thiếu like, bit 1: thiếu cmt, bit 2: thiếu tag, bit 3: thiếu share)
		def stat(mem):
			ret = 0
			if self.app.chk_bump_react.isChecked() and mem.get("React", 0) == 0: ret |= 1
			if self.app.chk_bump_cmt.isChecked() and mem.get("Comment", 0) < cmt_min: ret |= 2
			if self.app.chk_bump_mention.isChecked() and mem.get("Tag", 0) < mnt_min: ret |= 4
			if self.app.chk_bump_share.isChecked() and mem.get("Share", 0) == 0: ret |= 8
			return ret

		# hàm check xem có thực sự là không bump không
		def nobump(mem):
			if (self.app.chk_bump_react.isChecked() != True or mem.get("React", 0) == 0) and (self.app.chk_bump_cmt.isChecked() != True or mem.get("Comment", 0) == 0) and (self.app.chk_bump_mention.isChecked() != True or mem.get("Tag", 0) == 0) and (self.app.chk_bump_share.isChecked() != True or mem.get("Share", 0) == 0): return True
			return False

		# hàm xuất trạng thái (dạng text)
		def stat_txt(mem):
			t = stat(mem)
			if mem["UID"] == "": return "Thiếu UID"
			if t == 0: return "Đã bump đủ"
			if nobump(mem): return "Không bump"
			return "Chưa bump đủ"

		# hàm xuất phần ghi chú
		def notes(mem):
			t = stat(mem)
			if t != 0 and nobump(mem) == False:
				ret = "Thiếu "
				if t & 1:
					ret += "react"
					if t & 2 or t & 4: ret += ", "
				if t & 2:
					ret += str(cmt_min - mem.get("Comment", 0)) + " cmt"
					if t & 4 or t & 8: ret += ", "
				if t & 4:
					ret += str(mnt_min - mem.get("Tag", 0)) + " tag"
					if t & 8: ret += ", "
				if t & 8: ret += "share"
				return ret
			else: return ""
		
		self.app.wid_state_sig.emit(False)

		tstart = time.time() # đo thời gian check

		# lấy ID post
		self.app.bump_stat_sig.emit(0, "Đang lấy các thông tin về post...")
		page_id, post_id, author = get_ids("https://m.facebook.com/" + post_url, self.app)
		# print(f"{page_id} {post_id} {author}")
		global acc_pol_blist, acc_pol_wlist
		if page_id in acc_pol_blist or (len(acc_pol_wlist) != 0 and not page_id in acc_pol_wlist):
			self.app.msgbox("warning", "Thông báo", "Chính sách của đơn vị cấp tài khoản không cho phép check trang này.\nNếu bạn nghĩ đây là một sự nhầm lẫn, hãy liên hệ đơn vị cấp tài khoản của bạn.")
			self.app.wid_state_sig.emit(True)
			self.finished.emit(); return

		# check comment
		if self.app.chk_bump_cmt.isChecked():
			self.app.bump_stat_sig.emit(0, "Đang check comment...")
			check_cmt(members, post_id, self.app.chk_bump_reply.isChecked(), self.app.chk_bump_mention.isChecked(), self.app.chk_bump_mntexc.isChecked(), self.app.chk_bump_aux.isChecked() and self.app.chk_bump_aux.isEnabled(), self.app.bump_stat_sig, self.app)

		# check react
		if self.app.chk_bump_react.isChecked():
			self.app.bump_stat_sig.emit(33, "Đang check react...")
			check_react(members, post_id, author, self.app.bump_stat_sig, 33, self.app)

		# check share
		if self.app.chk_bump_share.isChecked():
			self.app.bump_stat_sig.emit(66, "Đang check share...")
			check_share(members, post_id, author, self.app.bump_stat_sig, 66, self.app)
		
		# lấy cap
		if self.app.chk_bump_scap.isChecked():
			self.app.bump_stat_sig.emit(83, "Đang lấy cap share...")
			get_scap(members, post_id, self.app.bump_stat_sig, self.app)

		self.app.bump_stat_sig.emit(99, "Đang ghi kết quả bump vào file...")
		for mem in members:
			if author in mem["UID"]:
				mem["Trạng thái"] = "Đã bump đủ"
				mem["Ghi chú"] = "Người đăng post"
			else:
				mem["Trạng thái"] = stat_txt(mem)
				mem["Ghi chú"] = notes(mem)
			mem["React"] = mem.get("React", 0)
			mem["Comment"] = mem.get("Comment", 0)
			mem["Tag"] = mem.get("Tag", 0)
			mem["Share"] = mem.get("Share", 0)
		if self.app.chk_bump_react.isChecked(): title["React"] = "React"
		if self.app.chk_bump_cmt.isChecked(): title["Comment"] = "Comment"
		if self.app.chk_bump_mention.isChecked(): title["Tag"] = "Tag"
		if self.app.chk_bump_share.isChecked(): title["Share"] = "Share"
		if self.app.chk_bump_scap.isChecked(): title["Cap share"] = "Cap share"
		title.update({"Trạng thái": "Trạng thái", "Ghi chú": "Ghi chú"})
		if self.app.chk_bump_cmt_text.isChecked(): title["Các comment"] = "cmt_*"
		stats = None
		if self.app.chk_bump_stats.isChecked():
			stats = []
			if self.app.chk_bump_react.isChecked():
				stats.append(["Tổng số react:", str(total_react), "trong đó của thành viên:", str(mem_react), "=\"({}%)\"".format(round(100 * mem_react / total_react, 2))])
			if self.app.chk_bump_cmt.isChecked():
				stats.append(["Tổng số cmt:", str(total_cmt), "trong đó của thành viên:", str(mem_cmt), "=\"({}%)\"".format(round(100 * mem_cmt / total_cmt, 2))])
			if self.app.chk_bump_share.isChecked():
				stats.append(["Tổng số share:", str(total_share), "trong đó của thành viên:", str(mem_share), "=\"({}%)\"".format(round(100 * mem_share / total_share, 2))])
		write_file(output_fname, datetime.now().strftime("KẾT QUẢ CHECK BUMP BÀI {} LÚC %H:%M NGÀY %d-%m-%Y".format(post_id).encode('unicode-escape').decode()).encode().decode("unicode-escape"), title, members, stats)

		global uid_dict
		uid_dict = {}

		self.app.wid_state_sig.emit(True)

		self.app.bump_stat_sig.emit(100, "Check bump hoàn tất ({})\nThời gian thực hiện: {} giây".format(self.app.lne_bump_outfile.text(), time.time() - tstart))
		self.app.msgbox("information", "Thông báo", "Check bump hoàn tất. Kết quả được lưu vào file {} (trong folder {}).".format(os.path.basename(self.app.lne_bump_outfile.text()), os.path.dirname(self.app.lne_bump_outfile.text())))
		self.app.lne_bump_outfile.setText(os.path.join(save_dir, datetime.now().strftime("bump-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx")))) # đổi tên file mặc định

		self.finished.emit(); return

class chkcf_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(chkcf_worker, self).__init__()
		self.app = app

	def run(self):
		check_acc(self.app)

		members_fname = self.app.lne_cf_memfile.text().replace("\\", "/").replace("\"", "")
		output_fname = self.app.lne_cf_outfile.text().replace("\\", "/").replace("\"", "")
		post_url = re.sub("^.*com\/", "", self.app.lne_cf_post.text())

		# check các trường tên và link
		if members_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập file thành viên!")
			self.finished.emit(); return
		if output_fname == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập file kết quả!")
			self.finished.emit(); return
		if post_url == "":
			self.app.msgbox("warning", "Thông báo", "Chưa nhập link post!")
			self.finished.emit(); return

		# hiển thị cảnh báo nếu check confirm
		if self.app.chk_cf_poll.isChecked():
			if self.app.msgbox("question", "Thông báo", "Do lỗi hiển thị của Facebook bản mobile nên kết quả check poll sẽ không chính xác cho các lựa chọn trên 30 người chọn.\nTiếp tục check poll?") == QMessageBox.No:
				self.app.chk_cf_poll.setChecked(False)

		# load file thành viên
		self.app.bump_stat_sig.emit(0, "Đang đọc và xử lý file thành viên...")
		members = preprocess_mlist(read_file(members_fname), self.app)
		title = {}
		for t in members[0].keys():
			if type(t) != str: continue
			title[t] = t
		
		# hàm xuất trạng thái cf (bit 0: thiếu like, bit 1: thiếu cmt, bit 2: thiếu poll)
		def stat(mem):
			ret = 0
			if self.app.chk_cf_react.isChecked() and mem.get("React", 0) == 0: ret |= 1
			if self.app.chk_cf_cmt.isChecked() and mem.get("Comment", 0) == 0: ret |= 2
			if self.app.chk_cf_poll.isChecked() and len(poll_name) != 0 and mem.get("Poll", 0) == 0: ret |= 4
			return ret

		# hàm check xem có thực sự là không cf không
		def nocf(mem):
			if (self.app.chk_cf_react.isChecked() != True or mem.get("React", 0) == 0) and (self.app.chk_cf_cmt.isChecked() != True or mem.get("Comment", 0) == 0) and (self.app.chk_cf_poll.isChecked() != True or len(poll_name) == 0 or (len(poll_name) != 0 and mem.get("poll", 0) == 0)): return True
			return False

		# hàm xuất trạng thái (dạng text)
		def stat_txt(mem):
			t = stat(mem)
			if mem["UID"] == "": return "Thiếu UID"
			if t == 0: return "Đã cf"
			if nocf(mem): return "Không cf"
			return "Chưa cf đủ"

		# hàm xuất phần ghi chú
		def notes(mem):
			t = stat(mem)
			if t != 0 and nocf(mem) == False:
				ret = "Thiếu "
				if t & 1:
					ret += "react"
					if t & 2 or t & 4: ret += ", "
				if t & 2:
					ret += "cmt"
					if t & 4: ret += ", "
				if t & 4: ret += "tick poll"
				return ret
			else: return ""

		self.app.wid_state_sig.emit(False)

		tstart = time.time() # đo thời gian check

		# lấy ID post
		self.app.cf_stat_sig.emit(0, "Đang lấy ID post và UID người đăng post.")
		page_id, post_id, author = get_ids("https://m.facebook.com/" + post_url, self.app)
		global acc_pol_blist, acc_pol_wlist
		if page_id in acc_pol_blist or (len(acc_pol_wlist) != 0 and not page_id in acc_pol_wlist):
			self.app.msgbox("warning", "Thông báo", "Chính sách của đơn vị cấp tài khoản không cho phép check trang này.\nNếu bạn nghĩ đây là một sự nhầm lẫn, hãy liên hệ đơn vị cấp tài khoản của bạn.")
			self.app.wid_state_sig.emit(True)
			self.finished.emit(); return

		# check comment
		if self.app.chk_cf_cmt.isChecked():
			self.app.cf_stat_sig.emit(0, "Đang check comment...")
			check_cmt(members, post_id, True, False, False, self.app.chk_cf_aux.isChecked() and self.app.chk_cf_aux.isEnabled(), self.app.cf_stat_sig, self.app)

		# check react
		if self.app.chk_cf_react.isChecked():
			self.app.cf_stat_sig.emit(33, "Đang check react...")
			check_react(members, post_id, author, self.app.cf_stat_sig, 33)

		# check poll
		if self.app.chk_cf_poll.isChecked():
			self.app.cf_stat_sig.emit(66, "Đang check poll...")
			check_poll(members, post_id, self.app.cf_stat_sig, self.app, 66)

		# print(members)

		self.app.cf_stat_sig.emit(99, "Đang ghi kết quả confirm vào file.")
		for mem in members:
			if author in mem["UID"]:
				mem["Trạng thái"] = "Đã cf"
				mem["Ghi chú"] = "Người đăng post"
			else:
				mem["Trạng thái"] = stat_txt(mem)
				mem["Ghi chú"] = notes(mem)
			# mem["UID"] = "=\"" + mem["UID"] + "\""
		
		if self.app.chk_cf_react.isChecked(): title["React"] = "React"
		if self.app.chk_cf_cmt.isChecked(): title["Comment"] = "Comment"
		if self.app.chk_cf_poll.isChecked() and len(poll_name) == 0: title["Poll"] = "Poll"
		for id in poll_name: title["[Poll] " + poll_name[id]] = "poll_" + str(id)
		title.update({"Trạng thái": "Trạng thái", "Ghi chú": "Ghi chú"})
		if self.app.chk_cf_cmt.isChecked(): title["Các comment"] = "cmt_*"
		write_file(output_fname, datetime.now().strftime("KẾT QUẢ CHECK CF BÀI {} LÚC %H:%M NGÀY %d-%m-%Y".format(post_id).encode('unicode-escape').decode()).encode().decode("unicode-escape"), title, members)

		global uid_dict
		uid_dict = {}

		self.app.wid_state_sig.emit(True)

		self.app.cf_stat_sig.emit(100, "Check confirm hoàn tất ({})\nThời gian thực hiện: {} giây".format(self.app.lne_cf_outfile.text(), time.time() - tstart))
		self.app.msgbox("information", "Thông báo", "Check confirm hoàn tất. Kết quả được lưu vào file {} (trong folder {}).".format(os.path.basename(self.app.lne_cf_outfile.text()), os.path.dirname(self.app.lne_cf_outfile.text())))
		self.app.lne_bump_outfile.setText(os.path.join(save_dir, datetime.now().strftime("cf-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx")))) # đổi tên file mặc định

class trial_worker(QtCore.QObject):
	finished = QtCore.pyqtSignal() # gọi khi hoàn tất

	def __init__(self, app):
		super(trial_worker, self).__init__()
		self.app = app

	def run(self):
		self.app.wid_state_sig.emit(False)
		# lấy UUID máy
		if platform.system() == "Windows":
			smuuid = subprocess.run("wmic csproduct get UUID".split(), capture_output=True).stdout.decode("ascii").replace("UUID", "").replace("\r", "").replace("\n", "").replace(" ", "")
		if smuuid == "FFFFFFFF-FFFF-FFFF-FFFF-FFFFFFFFFFFF":
			self.app.msgbox("warning", "Thông báo", "Máy tính của bạn không đủ điều kiện để dùng thử phần mềm.")
			self.app.wid_state_sig.emit(True)
			self.finished.emit(); return
		resp = requests.post("https://itsmevjnk.mooo.com/hrng/trial.php", data = {"uuid": smuuid}, verify=not py_legacy).text
		if resp == "ERR":
			self.app.msgbox("warning", "Thông báo", "Không thể kết nối với server tại thời điểm này. Vui lòng thử lại sau.")
			self.app.wid_state_sig.emit(True)
			self.finished.emit(); return
		else:
			self.app.lne_email.setText(smuuid.lower().replace("-", "") + "@trial.hrng.itsmevjnk.ga")
			self.app.lne_password.setText(smuuid + "_trial_do_not_steal")
			self.app.wid_state_sig.emit(True)
			self.login_worker = login_worker(self.app)
			self.login_thread = QtCore.QThread()
			self.login_worker.moveToThread(self.login_thread)
			self.login_thread.started.connect(self.login_worker.run)
			self.login_worker.finished.connect(self.login_thread.quit)
			self.login_worker.finished.connect(self.login_worker.deleteLater)
			self.login_thread.finished.connect(self.login_thread.deleteLater)
			self.login_thread.start()

class GUI(QMainWindow):
	msgbox_sig = QtCore.pyqtSignal(str, str, str, str) # nhận tín hiệu mở message box từ các thread
	msgbox_resp = {} # chứa kết quả trả về thread từ những lần mở message box

	init_stat_sig = QtCore.pyqtSignal(int, str)
	bump_stat_sig = QtCore.pyqtSignal(int, str)
	cf_stat_sig = QtCore.pyqtSignal(int, str)
	uid_stat_sig = QtCore.pyqtSignal(int, str, str)
	mlist_stat_sig = QtCore.pyqtSignal(int, str)
	switch_wid_sig = QtCore.pyqtSignal(int)
	exit_sig = QtCore.pyqtSignal()
	current_wid = 0
	wid_state_sig = QtCore.pyqtSignal(bool)

	def __init__(self):
		# khởi tạo cửa sổ
		QMainWindow.__init__(self)
		self.setMinimumSize(QtCore.QSize(64, 32))
		self.setWindowTitle("{} r{}".format(sw_name, sw_rel))

		# khởi tạo widget chứa các widget khác
		self.wid_main = QStackedWidget(self)
		self.setCentralWidget(self.wid_main)

		# tạo các widget khác có liên quan
		self.init_login()
		self.init_init()
		self.init_tabs()
		
		# thiết lập các signal
		self.msgbox_sig.connect(self.do_show_msgbox)
		self.switch_wid_sig.connect(self.switch_widget)
		self.exit_sig.connect(self.do_exit)
		self.wid_state_sig.connect(self.do_wid_state)

		self.switch_widget(0) # resize lại cửa sổ login
		self.show() # hiển thị cửa sổ khi đã sẵn sàng

		if py_legacy: QMessageBox.warning(self, "Thông báo", "Bạn đang sử dụng HRng trên một phiên bản hệ điều hành cũ. Một số tính năng sẽ bị vô hiệu hoá để đảm bảo phần mềm hoạt động đúng.\nHãy nâng cấp hệ điều hành để có trải nghiệm sử dụng tốt nhất.")
	
	# hàm xử lý tín hiệu mở message box
	def do_show_msgbox(self, tname, type, title, text):
		self.msgbox_resp[tname] = getattr(QMessageBox, type)(self, title, text)

	# chuyển widget hiển thị
	def switch_widget(self, idx): # 0: login, 1: init, 2: tabs
		self.wid_main.setCurrentIndex(idx)
		self.resize(self.wid_main.currentWidget().minimumSizeHint())
		self.current_wid = idx # báo hiệu cho thread
	
	# hiển thị hộp thoại
	def msgbox(self, type, title, text):
		tname = current_thread().name
		self.msgbox_resp[tname] = None
		self.msgbox_sig.emit(tname, type, title, text)
		while self.msgbox_resp[tname] == None: time.sleep(0.1)
		return self.msgbox_resp[tname]

	# đặt trạng thái widget
	def do_wid_state(self, state):
		self.wid_main.setEnabled(state)

	# thoát chương trình
	def do_exit(self):
		# đóng trình duyệt
		try: driver.quit()
		except: pass
		sys.exit(0)
	
	def init_login(self):
		wid_login = QWidget()
		lyt_login = QVBoxLayout()
		wid_login.setLayout(lyt_login)
		self.wid_main.addWidget(wid_login)

		lyt_login_grid = QGridLayout()
		lyt_login.addLayout(lyt_login_grid)
		self.lne_email = QLineEdit()
		self.lne_email.setText(config.get("email", ""))
		lyt_login_grid.addWidget(QLabel("Email:                                  "), 0, 0)
		lyt_login_grid.addWidget(self.lne_email, 1, 0)
		self.lne_password = QLineEdit()
		self.lne_password.setEchoMode(QLineEdit.Password)
		lyt_login_grid.addWidget(QLabel("Mật khẩu:                               "), 0, 1)
		lyt_login_grid.addWidget(self.lne_password, 1, 1)

		lyt_cookies = QHBoxLayout()
		lyt_login.addLayout(lyt_cookies)
		self.lne_cookies = QLineEdit()
		self.lne_cookies.setText(config.get("cookies", ""))
		lyt_cookies.addWidget(QLabel("Cookies FB:"))
		lyt_cookies.addWidget(self.lne_cookies)
		btn_import_cookies = QPushButton("Nhập")
		lyt_cookies.addWidget(btn_import_cookies, 0)

		self.chk_cookies_aux = QCheckBox("Sử dụng tài khoản khác để check comment và tag:")
		lyt_login.addWidget(self.chk_cookies_aux)

		lyt_cookies_aux = QHBoxLayout()
		lyt_login.addLayout(lyt_cookies_aux)
		self.lne_cookies_aux = QLineEdit()
		self.lne_cookies_aux.setText(config.get("cookies_aux", ""))
		self.lne_cookies_aux.setEnabled(False)
		lyt_cookies_aux.addWidget(self.lne_cookies_aux)
		self.btn_import_cookies_aux = QPushButton("Nhập")
		self.btn_import_cookies_aux.setEnabled(False)
		lyt_cookies_aux.addWidget(self.btn_import_cookies_aux, 0)

		self.chk_firefox = QCheckBox("Sử dụng Firefox thay cho Chromium")
		lyt_login.addWidget(self.chk_firefox)

		lyt_login_opts = QHBoxLayout()
		lyt_login.addLayout(lyt_login_opts)
		btn_init = QPushButton("Đăng nhập")
		btn_init.setDefault(True)
		lyt_login_opts.addWidget(btn_init, 1)
		btn_trial = QPushButton("Dùng thử")
		lyt_login_opts.addWidget(btn_trial, 1)

		btn_init.clicked.connect(self.do_login)
		btn_trial.clicked.connect(self.do_trial)
		btn_import_cookies.clicked.connect(self.do_import_cookies)

		self.chk_cookies_aux.stateChanged.connect(self.cookies_aux_chkbox)
		self.btn_import_cookies_aux.clicked.connect(self.do_import_cookies_aux)

	def import_cookies(self, fname):
		cookies_str = ""
		with open(fname, "r") as f:
			cookies_txt = f.read()
			for line in cookies_txt.splitlines():
				line_split = line.split("\t")
				if len(line_split) == 7:
					cookies_str += "{}={};".format(line_split[-2], line_split[-1])
		return cookies_str

	def do_import_cookies(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file cookies", "", "File text (*.txt);;Tất cả file (*)", options = QFileDialog.Options())
		if fname:
			self.lne_cookies.setText(self.import_cookies(fname))
	
	def do_import_cookies_aux(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file cookies", "", "File text (*.txt);;Tất cả file (*)", options = QFileDialog.Options())
		if fname:
			self.lne_cookies_aux.setText(self.import_cookies(fname))

	def do_login(self):
		self.login_worker = login_worker(self)
		self.login_thread = QtCore.QThread()
		self.login_worker.moveToThread(self.login_thread)
		self.login_thread.started.connect(self.login_worker.run)
		self.login_worker.finished.connect(self.login_thread.quit)
		self.login_worker.finished.connect(self.login_worker.deleteLater)
		self.login_thread.finished.connect(self.login_thread.deleteLater)
		self.login_thread.start()
	
	def do_trial(self):
		self.trial_worker = trial_worker(self)
		self.trial_thread = QtCore.QThread()
		self.trial_worker.moveToThread(self.trial_thread)
		self.trial_thread.started.connect(self.trial_worker.run)
		self.trial_worker.finished.connect(self.trial_thread.quit)
		self.trial_worker.finished.connect(self.trial_worker.deleteLater)
		self.trial_thread.finished.connect(self.trial_thread.deleteLater)
		self.trial_thread.start()

	def init_init(self):
		wid_init = QWidget()
		lyt_init = QVBoxLayout()
		wid_init.setLayout(lyt_init)
		self.wid_main.addWidget(wid_init)

		self.lbl_init_status = QLabel("Đang khởi động...                                                    ")
		lyt_init.addWidget(self.lbl_init_status)
		self.pbr_init = QProgressBar()
		lyt_init.addWidget(self.pbr_init)
		
		self.init_stat_sig.connect(self.do_init_stat)

	def do_init_stat(self, v, s):
		if s != ' ': self.lbl_init_status.setText(s)
		if v != -1: self.pbr_init.setValue(v)
	
	def init_tabs(self):
		wid_tabs = QTabWidget()
		self.wid_main.addWidget(wid_tabs)

		# tab check bump
		def init_chkbump(self):
			wid_chkbump = QWidget()
			lyt_chkbump = QVBoxLayout()
			wid_chkbump.setLayout(lyt_chkbump)
			wid_tabs.addTab(wid_chkbump, "Check &bump")

			# layout phần chọn file thành viên
			lyt_bump_memfile = QHBoxLayout()
			lyt_chkbump.addLayout(lyt_bump_memfile)
			lyt_bump_memfile.addWidget(QLabel("File thành viên:"), 0)
			self.lne_bump_memfile = QLineEdit()
			lyt_bump_memfile.addWidget(self.lne_bump_memfile, 1)
			btn_bump_memfile_browse = QPushButton("Tìm")
			lyt_bump_memfile.addWidget(btn_bump_memfile_browse, 0)

			# layout phần điền link post
			lyt_bump_post = QHBoxLayout()
			lyt_chkbump.addLayout(lyt_bump_post)
			lyt_bump_post.addWidget(QLabel("Link bài:"), 0)
			self.lne_bump_post = QLineEdit()
			lyt_bump_post.addWidget(self.lne_bump_post, 1)

			# layout phần chọn thông số check bump
			lyt_bump_options = QGridLayout()
			lyt_chkbump.addLayout(lyt_bump_options)
			self.chk_bump_react = QCheckBox("Bắt buộc react")
			self.chk_bump_react.setChecked(True)
			lyt_bump_options.addWidget(self.chk_bump_react, 0, 0)
			self.chk_bump_cmt = QCheckBox("Bắt buộc comment đủ")
			self.chk_bump_cmt.setChecked(True)
			lyt_bump_options.addWidget(self.chk_bump_cmt, 0, 1)
			self.chk_bump_mention = QCheckBox("Bắt buộc tag đủ")
			self.chk_bump_mention.setChecked(False)
			lyt_bump_options.addWidget(self.chk_bump_mention, 0, 2)
			self.chk_bump_share = QCheckBox("Bắt buộc share")
			self.chk_bump_share.setChecked(True)
			lyt_bump_options.addWidget(self.chk_bump_share, 0, 3)
			lyt_bump_options.addWidget(QLabel("Số comment tối thiểu:"), 1, 0)
			self.sbx_bump_cmtmin = QSpinBox()
			self.sbx_bump_cmtmin.setValue(3)
			lyt_bump_options.addWidget(self.sbx_bump_cmtmin, 1, 1)
			lyt_bump_options.addWidget(QLabel("Số người tag tối thiểu:"), 2, 0)
			self.sbx_bump_mntmin = QSpinBox()
			self.sbx_bump_mntmin.setValue(3)
			self.sbx_bump_mntmin.setEnabled(False)
			lyt_bump_options.addWidget(self.sbx_bump_mntmin, 2, 1)
			self.chk_bump_reply = QCheckBox("Check cả phản hồi bình luận")
			self.chk_bump_reply.setChecked(False)
			lyt_bump_options.addWidget(self.chk_bump_reply, 1, 2, 1, 2)
			self.chk_bump_mntexc = QCheckBox("Bỏ qua tag các thành viên")
			self.chk_bump_mntexc.setChecked(False)
			self.chk_bump_mntexc.setEnabled(False)
			lyt_bump_options.addWidget(self.chk_bump_mntexc, 2, 2, 1, 2)
			self.chk_bump_cmt_text = QCheckBox("Xuất nội dung bình luận")
			self.chk_bump_cmt_text.setChecked(True)
			lyt_bump_options.addWidget(self.chk_bump_cmt_text, 3, 0, 1, 2)
			self.chk_bump_stats = QCheckBox("Xuất thống kê chi tiết")
			self.chk_bump_stats.setChecked(False)
			lyt_bump_options.addWidget(self.chk_bump_stats, 3, 2, 1, 2)
			self.chk_bump_aux = QCheckBox("Sử dụng tài khoản riêng để check comment và tag")
			self.chk_bump_aux.setChecked(False)
			self.chk_bump_aux.setEnabled(False)
			lyt_bump_options.addWidget(self.chk_bump_aux, 4, 0, 1, 2)
			self.chk_bump_scap = QCheckBox("Lấy cap bài share")
			self.chk_bump_scap.setChecked(False)
			lyt_bump_options.addWidget(self.chk_bump_scap, 4, 2, 1, 1)

			# layout phần chọn file kết quả
			lyt_bump_outfile = QHBoxLayout()
			lyt_chkbump.addLayout(lyt_bump_outfile)
			lyt_bump_outfile.addWidget(QLabel("File kết quả:"), 0)
			self.lne_bump_outfile = QLineEdit()
			self.lne_bump_outfile.setText(os.path.join(save_dir, datetime.now().strftime("bump-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
			lyt_bump_outfile.addWidget(self.lne_bump_outfile, 1)
			btn_bump_outfile_browse = QPushButton("Tìm")
			lyt_bump_outfile.addWidget(btn_bump_outfile_browse, 0)

			# phần các nút chọn và trạng thái
			btn_bump_start = QPushButton("Check bump")
			btn_bump_start.setDefault(True)
			lyt_chkbump.addWidget(btn_bump_start)
			self.pbr_bump = QProgressBar()
			lyt_chkbump.addWidget(self.pbr_bump)
			self.lbl_bump_status = QLabel("Bấm Check bump để bắt đầu...\n")
			lyt_chkbump.addWidget(self.lbl_bump_status)

			lyt_chkbump.addStretch()

			self.bump_stat_sig.connect(self.do_bump_stat)
			btn_bump_start.clicked.connect(self.do_chkbump)

			btn_bump_memfile_browse.clicked.connect(self.bump_browse_memfile)
			btn_bump_outfile_browse.clicked.connect(self.bump_browse_outfile)

			self.chk_bump_cmt.stateChanged.connect(self.bump_cmt_chkbox)
			self.chk_bump_mention.stateChanged.connect(self.bump_mention_chkbox)
			self.chk_bump_share.stateChanged.connect(self.bump_share_chkbox)
		init_chkbump(self)

		# tab check cf
		def init_chkcf(self):
			wid_chkcf = QWidget()
			lyt_chkcf = QVBoxLayout()
			wid_chkcf.setLayout(lyt_chkcf)
			wid_tabs.addTab(wid_chkcf, "Check &confirm")

			# layout phần chọn file thành viên
			lyt_cf_memfile = QHBoxLayout()
			lyt_chkcf.addLayout(lyt_cf_memfile)
			lyt_cf_memfile.addWidget(QLabel("File thành viên:"), 0)
			self.lne_cf_memfile = QLineEdit()
			lyt_cf_memfile.addWidget(self.lne_cf_memfile, 1)
			btn_cf_memfile_browse = QPushButton("Tìm")
			lyt_cf_memfile.addWidget(btn_cf_memfile_browse, 0)

			# layout phần điền link post
			lyt_cf_post = QHBoxLayout()
			lyt_chkcf.addLayout(lyt_cf_post)
			lyt_cf_post.addWidget(QLabel("Link bài:"), 0)
			self.lne_cf_post = QLineEdit()
			lyt_cf_post.addWidget(self.lne_cf_post, 1)

			# layout phần chọn thông số check cf
			lyt_cf_options = QGridLayout()
			lyt_chkcf.addLayout(lyt_cf_options)
			self.chk_cf_react = QCheckBox("Bắt buộc react")
			self.chk_cf_react.setChecked(False)
			lyt_cf_options.addWidget(self.chk_cf_react, 0, 0)
			self.chk_cf_cmt = QCheckBox("Bắt buộc comment")
			self.chk_cf_cmt.setChecked(True)
			lyt_cf_options.addWidget(self.chk_cf_cmt, 0, 1)
			self.chk_cf_poll = QCheckBox("Bắt buộc tick poll (nếu có)")
			self.chk_cf_poll.setChecked(True)
			lyt_cf_options.addWidget(self.chk_cf_poll, 0, 2)
			self.chk_cf_cmt_text = QCheckBox("Xuất nội dung bình luận")
			self.chk_cf_cmt_text.setChecked(True)
			lyt_cf_options.addWidget(self.chk_cf_cmt_text, 1, 0, 1, 2)
			self.chk_cf_aux = QCheckBox("Sử dụng tài khoản riêng để check comment")
			self.chk_cf_aux.setChecked(False)
			self.chk_cf_aux.setEnabled(False)
			lyt_cf_options.addWidget(self.chk_cf_aux, 4, 0, 1, 3)

			# layout phần chọn file kết quả
			lyt_cf_outfile = QHBoxLayout()
			lyt_chkcf.addLayout(lyt_cf_outfile)
			lyt_cf_outfile.addWidget(QLabel("File kết quả:"), 0)
			self.lne_cf_outfile = QLineEdit()
			self.lne_cf_outfile.setText(os.path.join(save_dir, datetime.now().strftime("cf-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
			lyt_cf_outfile.addWidget(self.lne_cf_outfile, 1)
			btn_cf_outfile_browse = QPushButton("Tìm")
			lyt_cf_outfile.addWidget(btn_cf_outfile_browse, 0)

			# phần các nút chọn và trạng thái
			btn_cf_start = QPushButton("Check confirm")
			btn_cf_start.setDefault(True)
			lyt_chkcf.addWidget(btn_cf_start)
			self.pbr_cf = QProgressBar()
			lyt_chkcf.addWidget(self.pbr_cf)
			self.lbl_cf_status = QLabel("Bấm Check confirm để bắt đầu...\n")
			lyt_chkcf.addWidget(self.lbl_cf_status)

			lyt_chkcf.addStretch()

			self.cf_stat_sig.connect(self.do_cf_stat)
			btn_cf_start.clicked.connect(self.do_chkcf)

			btn_cf_memfile_browse.clicked.connect(self.cf_browse_memfile)
			btn_cf_outfile_browse.clicked.connect(self.cf_browse_outfile)

			self.chk_cf_cmt.stateChanged.connect(self.cf_cmt_chkbox)
		init_chkcf(self)

		# tab lấy UID
		def init_uid(self):
			wid_uid = QWidget()
			lyt_uid = QVBoxLayout()
			wid_uid.setLayout(lyt_uid)
			wid_tabs.addTab(wid_uid, "Tìm &UID")

			# phần link fb
			lyt_uid_link = QHBoxLayout()
			lyt_uid.addLayout(lyt_uid_link)
			lyt_uid_link.addWidget(QLabel("Link tài khoản:"), 0)
			self.lne_uid_link = QLineEdit()
			lyt_uid_link.addWidget(self.lne_uid_link, 1)

			# phần kết quả
			lyt_uid_out = QHBoxLayout()
			lyt_uid.addLayout(lyt_uid_out)
			lyt_uid_out.addWidget(QLabel("UID:"), 0)
			self.lne_uid = QLineEdit()
			self.lne_uid.setEnabled(False)
			self.lne_uid.setAlignment(QtCore.Qt.AlignCenter)
			lyt_uid_out.addWidget(self.lne_uid, 1)
			self.btn_uid_copy = QPushButton("Copy")
			self.btn_uid_copy.setEnabled(False)
			lyt_uid_out.addWidget(self.btn_uid_copy, 0)

			# lấy UID nhiều tài khoản
			# phần input
			lyt_uid_batch_in = QHBoxLayout()
			lyt_uid.addLayout(lyt_uid_batch_in)
			lyt_uid_batch_in.addWidget(QLabel("File danh sách link:"), 0)
			self.lne_uid_in = QLineEdit()
			lyt_uid_batch_in.addWidget(self.lne_uid_in, 1)
			btn_uid_in_browse = QPushButton("Tìm")
			lyt_uid_batch_in.addWidget(btn_uid_in_browse, 0)

			# phần output
			lyt_uid_batch_out = QHBoxLayout()
			lyt_uid.addLayout(lyt_uid_batch_out)
			lyt_uid_batch_out.addWidget(QLabel("File kết quả:"), 0)
			self.lne_uid_out = QLineEdit()
			self.lne_uid_out.setText(os.path.join(save_dir, datetime.now().strftime("uid-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
			lyt_uid_batch_out.addWidget(self.lne_uid_out, 1)
			btn_uid_out_browse = QPushButton("Tìm")
			lyt_uid_batch_out.addWidget(btn_uid_out_browse, 0)

			# phần nút bắt đầu
			btn_uid = QPushButton("Tìm UID")
			btn_uid.setDefault(True)
			lyt_uid.addWidget(btn_uid)
			
			# phần trạng thái
			self.lbl_uid_stat = QLabel()
			lyt_uid.addWidget(self.lbl_uid_stat)
			self.pbr_uid = QProgressBar()
			lyt_uid.addWidget(self.pbr_uid)

			lyt_uid.addStretch()

			self.uid_stat_sig.connect(self.do_uid_stat)
			self.btn_uid_copy.clicked.connect(self.do_uid_copy)
			btn_uid.clicked.connect(self.do_getuid)

			btn_uid_in_browse.clicked.connect(self.uid_browse_infile)
			btn_uid_out_browse.clicked.connect(self.uid_browse_outfile)
		init_uid(self)

		# tab lấy danh sách thành viên
		def init_mlist(self):
			wid_mlist = QWidget()
			lyt_mlist = QVBoxLayout()
			wid_mlist.setLayout(lyt_mlist)
			wid_tabs.addTab(wid_mlist, "Lấy &danh sách thành viên")

			# phần link fb
			lyt_mlist_link = QHBoxLayout()
			lyt_mlist.addLayout(lyt_mlist_link)
			lyt_mlist_link.addWidget(QLabel("Link nhóm:"), 0)
			self.lne_mlist_link = QLineEdit()
			lyt_mlist_link.addWidget(self.lne_mlist_link, 1)

			# các tuỳ chọn
			self.chk_mlist_uid = QCheckBox("Lấy UID")
			lyt_mlist.addWidget(self.chk_mlist_uid)
			lyt_mlist_cmp = QHBoxLayout()
			lyt_mlist.addLayout(lyt_mlist_cmp)
			self.chk_mlist_cmp = QCheckBox("Đối chiếu:")
			lyt_mlist_cmp.addWidget(self.chk_mlist_cmp, 0)
			self.lne_mlist_cmp = QLineEdit()
			self.lne_mlist_cmp.setEnabled(False)
			lyt_mlist_cmp.addWidget(self.lne_mlist_cmp, 1)
			self.btn_mlist_cmp_browse = QPushButton("Tìm")
			self.btn_mlist_cmp_browse.setEnabled(False)
			lyt_mlist_cmp.addWidget(self.btn_mlist_cmp_browse, 0)

			# phần output
			lyt_mlist_out = QHBoxLayout()
			lyt_mlist.addLayout(lyt_mlist_out)
			lyt_mlist_out.addWidget(QLabel("File kết quả:"), 0)
			self.lne_mlist_out = QLineEdit()
			self.lne_mlist_out.setText(os.path.join(save_dir, datetime.now().strftime("mlist-%d%m%Y_%H%M%S.{}".format("csv" if py_legacy else "xlsx"))))
			lyt_mlist_out.addWidget(self.lne_mlist_out, 1)
			btn_mlist_out_browse = QPushButton("Tìm")
			lyt_mlist_out.addWidget(btn_mlist_out_browse, 0)

			# phần nút bắt đầu
			btn_mlist = QPushButton("Lấy danh sách")
			btn_mlist.setDefault(True)
			lyt_mlist.addWidget(btn_mlist)
			
			# phần trạng thái
			self.lbl_mlist_stat = QLabel()
			lyt_mlist.addWidget(self.lbl_mlist_stat)
			self.pbr_mlist = QProgressBar()
			lyt_mlist.addWidget(self.pbr_mlist)

			lyt_mlist.addStretch()

			self.mlist_stat_sig.connect(self.do_mlist_stat)
			btn_mlist.clicked.connect(self.do_mlist)

			self.btn_mlist_cmp_browse.clicked.connect(self.mlist_browse_cmpfile)
			btn_mlist_out_browse.clicked.connect(self.mlist_browse_outfile)

			self.chk_mlist_cmp.stateChanged.connect(self.mlist_cmp_chkbox)
		init_mlist(self)

		# tab thông tin phần mềm
		def init_about(self):
			wid_about = QWidget()
			lyt_about = QVBoxLayout()
			wid_about.setLayout(lyt_about)
			wid_tabs.addTab(wid_about, "&Về phần mềm")
			lbl_about = QLabel("{} r{}\nPhần mềm hỗ trợ quản lý nhân sự\nĐược phát triển bởi:\nNguyễn Thành Vinh (K113 Lý, THPT Chu Văn An)".format(sw_name, sw_rel))
			lbl_about.setAlignment(QtCore.Qt.AlignCenter)
			lyt_about.addWidget(lbl_about)
			self.lbl_version = QLabel()
			self.lbl_version.setAlignment(QtCore.Qt.AlignCenter)
			lyt_about.addWidget(self.lbl_version)
			self.lbl_license = QLabel()
			self.lbl_license.setAlignment(QtCore.Qt.AlignCenter)
			lyt_about.addWidget(self.lbl_license)
			lyt_about.addStretch()
		init_about(self)
	
	def do_getuid(self):
		self.getuid_worker = getuid_worker(self)
		self.getuid_thread = QtCore.QThread()
		self.getuid_worker.moveToThread(self.getuid_thread)
		self.getuid_thread.started.connect(self.getuid_worker.run)
		self.getuid_worker.finished.connect(self.getuid_thread.quit)
		self.getuid_worker.finished.connect(self.getuid_worker.deleteLater)
		self.getuid_thread.finished.connect(self.getuid_thread.deleteLater)
		self.getuid_thread.start() 
	def do_chkbump(self):
		self.chkbump_worker = chkbump_worker(self)
		self.chkbump_thread = QtCore.QThread()
		self.chkbump_worker.moveToThread(self.chkbump_thread)
		self.chkbump_thread.started.connect(self.chkbump_worker.run)
		self.chkbump_worker.finished.connect(self.chkbump_thread.quit)
		self.chkbump_worker.finished.connect(self.chkbump_worker.deleteLater)
		self.chkbump_thread.finished.connect(self.chkbump_thread.deleteLater)
		self.chkbump_thread.start()   
	def do_chkcf(self):
		self.chkcf_worker = chkcf_worker(self)
		self.chkcf_thread = QtCore.QThread()
		self.chkcf_worker.moveToThread(self.chkcf_thread)
		self.chkcf_thread.started.connect(self.chkcf_worker.run)
		self.chkcf_worker.finished.connect(self.chkcf_thread.quit)
		self.chkcf_worker.finished.connect(self.chkcf_worker.deleteLater)
		self.chkcf_thread.finished.connect(self.chkcf_thread.deleteLater)
		self.chkcf_thread.start()
	def do_mlist(self):
		self.mlist_worker = mlist_worker(self)
		self.mlist_thread = QtCore.QThread()
		self.mlist_worker.moveToThread(self.mlist_thread)
		self.mlist_thread.started.connect(self.mlist_worker.run)
		self.mlist_worker.finished.connect(self.mlist_thread.quit)
		self.mlist_worker.finished.connect(self.mlist_worker.deleteLater)
		self.mlist_thread.finished.connect(self.mlist_thread.deleteLater)
		self.mlist_thread.start()

	def do_bump_stat(self, v, s):
		if s != ' ': self.lbl_bump_status.setText(s)
		if v != -1: self.pbr_bump.setValue(v)
	def do_cf_stat(self, v, s):
		if s != ' ': self.lbl_cf_status.setText(s)
		if v != -1: self.pbr_cf.setValue(v)
	def do_uid_stat(self, v, uid, sta):
		if v != -1: self.pbr_uid.setValue(v)
		if uid != " ": self.lne_uid.setText(uid)
		if self.lne_uid.text().replace(" ", "") != "": self.btn_uid_copy.setEnabled(True)
		else: self.btn_uid_copy.setEnabled(False)
		if sta != " ": self.lbl_uid_stat.setText(sta)
	def do_mlist_stat(self, v, s) :
		if s != ' ': self.lbl_mlist_stat.setText(s)
		if v != -1: self.pbr_mlist.setValue(v)
	
	def bump_mention_chkbox(self):
		state = self.chk_bump_mention.isEnabled() and self.chk_bump_mention.isChecked()
		self.sbx_bump_mntmin.setEnabled(state)
		self.chk_bump_mntexc.setEnabled(state)
	
	def bump_cmt_chkbox(self):
		state = self.chk_bump_cmt.isChecked()
		self.chk_bump_mention.setEnabled(state)
		self.sbx_bump_cmtmin.setEnabled(state)
		self.chk_bump_reply.setEnabled(state)
		self.chk_bump_cmt_text.setEnabled(state)
		self.bump_mention_chkbox()
		if self.chk_cookies_aux.isChecked():
			self.chk_bump_aux.setEnabled(state)
			# self.chk_bump_aux.setChecked(state)

	def cf_cmt_chkbox(self):
		state = self.chk_cf_cmt.isChecked()
		self.chk_cf_cmt_text.setEnabled(state)
		if self.chk_cookies_aux.isChecked():
			self.chk_cf_aux.setEnabled(state)
			# self.chk_cf_aux.setChecked(state)
	
	def mlist_cmp_chkbox(self):
		state = self.chk_mlist_cmp.isChecked()
		self.lne_mlist_cmp.setEnabled(state)
		self.btn_mlist_cmp_browse.setEnabled(state)
	
	def bump_share_chkbox(self):
		self.chk_bump_scap.setEnabled(self.chk_bump_share.isChecked())
	
	def cookies_aux_chkbox(self):
		state = self.chk_cookies_aux.isChecked()
		self.lne_cookies_aux.setEnabled(state)
		self.btn_import_cookies_aux.setEnabled(state)
		self.chk_bump_aux.setEnabled(state)
		self.chk_bump_aux.setChecked(state)
		self.chk_cf_aux.setEnabled(state)

	def bump_browse_memfile(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file thành viên", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_bump_memfile.setText(fname.replace('/', os.path.sep))
	def bump_browse_outfile(self):
		fname, _ = QFileDialog.getSaveFileName(qt_win, "Chọn nơi lưu file kết quả check bump", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_bump_outfile.setText(fname.replace('/', os.path.sep))
	def cf_browse_memfile(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file thành viên", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_cf_memfile.setText(fname.replace('/', os.path.sep))
	def cf_browse_outfile(self):
		fname, _ = QFileDialog.getSaveFileName(qt_win, "Chọn nơi lưu file kết quả check confirm", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_cf_outfile.setText(fname.replace('/', os.path.sep))
	def uid_browse_infile(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file danh sách link Facebook", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_uid_in.setText(fname.replace('/', os.path.sep))
	def uid_browse_outfile(self):
		fname, _ = QFileDialog.getSaveFileName(qt_win, "Chọn nơi lưu file kết quả lấy UID", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_uid_out.setText(fname.replace('/', os.path.sep))
	def mlist_browse_cmpfile(self):
		fname, _ = QFileDialog.getOpenFileName(qt_win, "Chọn file danh sách link Facebook để đối chiếu", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_mlist_cmp.setText(fname.replace('/', os.path.sep))
	def mlist_browse_outfile(self):
		fname, _ = QFileDialog.getSaveFileName(qt_win, "Chọn nơi lưu file kết quả", "", "{}File CSV (*.csv);;Tất cả file (*)".format("" if py_legacy else "File Excel (*.xlsx);;"), options = QFileDialog.Options())
		if fname: self.lne_mlist_out.setText(fname.replace('/', os.path.sep))

	# xử lý yêu cầu thoát chương trình
	def closeEvent(self, event):
		if self.msgbox("question", "Thoát chương trình", "Bạn có muốn thoát khỏi chương trình không?") == QMessageBox.No:
			if not type(event) == bool: event.ignore()
		else:
			if not type(event) == bool: event.accept()
			else: self.do_exit()

	@QtCore.pyqtSlot()
	def do_uid_copy(self):
		uid = self.lne_uid.text()
		if uid != " ": pyperclip.copy(uid)

# exception hook, hi vọng nó bắt được cả exception ở các thread khác
def exchook(exctype, value, tb):
	fname = os.path.join(save_dir, datetime.now().strftime("bug-%d%m%Y_%H%M%S.txt"))
	with open(fname, "w", encoding="utf-8-sig") as file:
		traceback.print_exception(exctype, value, tb, file = file)
	qt_win.msgbox("critical", "Exception", "Đã có lỗi xảy ra trong quá trình chạy.\nVui lòng gửi file {} cho tác giả qua email itsmevjnk.work@gmail.com.\nBấm OK để kết thúc chương trình.".format(fname.replace('/', os.path.sep)))
	qt_win.exit_sig.emit()

# cython
def start():
	global qt_app, qt_win
	qt_app = QApplication(sys.argv)
	qt_win = GUI()
	sys.excepthook = exchook
	qt_app.exec_()
	qt_win.do_exit()

start()
