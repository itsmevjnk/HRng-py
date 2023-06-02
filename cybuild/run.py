#
#                       _oo0oo_
#                      o8888888o
#                      88" . "88
#                      (| -_- |)
#                      0\  =  /0
#                    ___/`---'\___
#                  .' \\|     |// '.
#                 / \\|||  :  |||// \
#                / _||||| -:- |||||- \
#               |   | \\\  -  /// |   |
#               | \_|  ''\---/''  |_/ |
#               \  .-\__  '-'  ___/-. /
#             ___'. .'  /--.--\  `. .'___
#          ."" '<  `.___\_<|>_/___.' >' "".
#         | | :  `- \`.;`\ _ /`;.`/ - ` : | |
#         \  \ `_.   \_ __\ /__ _/   .-` /  /
#     =====`-.____`.___ \_____/___.-`___.-'=====
#                       `=---='
#
#
#     ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#
#               佛祖保佑         永无BUG
#

import os
import re
import csv
import math
import requests
from requests.exceptions import Timeout
import time
from lxml.html import fromstring
from urllib.parse import unquote
import traceback
import json
import stat
import shutil
import platform
import zipfile
import tarfile
import pyperclip
import functools
import subprocess
import urllib3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from threading import current_thread
import pathlib
try: from subprocess import CREATE_NO_WINDOW
except: pass
import random

# mã hoá AES-256
from base64 import b64encode, b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from hashlib import sha256

# lấy thời gian chạy
from datetime import datetime, timezone
import dateparser
from dateutil.relativedelta import relativedelta

# GUI
from PyQt5.QtWidgets import *
from PyQt5 import QtCore

# selenium
import selenium

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.firefox.options import Options as FFOptions
from selenium.webdriver.firefox.service import Service as FFService

import bpexec